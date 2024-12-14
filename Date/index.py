import openpyxl
from datetime import datetime
from collections import defaultdict

# Charger le fichier Excel
wb = openpyxl.load_workbook('1x.xlsx')

# Sélectionner la feuille active (ou changez pour spécifier la feuille)
ws = wb.active

# Créer un dictionnaire pour stocker les matchs regroupés dans les onglets
favoris_domicile_moins_200 = []
favoris_domicile_plus_200 = []
favoris_exterieurs_moins_200 = []
favoris_exterieurs_plus_200 = []

# Fonction pour vérifier si une valeur est un nombre valide
def is_valid_number(value):
    try:
        # Tenter de convertir en entier
        int(value)
        return True
    except ValueError:
        # Retourner False si la conversion échoue
        return False

# Fonction pour convertir la date avec différents formats
def parse_date(date_str):
    if date_str is None:
        return None  # Si la valeur de date est None, on retourne None
    
    if isinstance(date_str, datetime):
        # Si c'est déjà un objet datetime, retourner tel quel
        return date_str
    
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):  # Essayer ces deux formats
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None  # Si la date ne correspond à aucun des formats

# Variables pour compter les matchs moins de 2 buts et plus de 2 buts
moins_de_2_buts = 0
plus_de_2_buts = 0

# Lire les données (en partant de la ligne 2 pour ignorer l'entête)
for row in ws.iter_rows(min_row=2, values_only=True):
    score = row[0]
    odds = row[1]  # 1XBET ODDS
    ou_2_5 = row[2]  # 1XBET O/U 2.5
    date = row[3]

    # Convertir la date
    parsed_date = parse_date(date)

    if odds and ou_2_5 and parsed_date:
        try:
            # Extraire les valeurs de cotes (séparer par '/')
            odds_values = odds.split('/')

            # Vérifier que nous avons bien 3 valeurs de cotes
            if len(odds_values) == 3:
                # Extraire la première et la dernière cote
                first_odds = int(odds_values[0])
                last_odds = int(odds_values[-1])

                # Trouver la plus petite cote
                smallest_odds = min(first_odds, last_odds)

                # Appliquer les règles pour les favoris domicile et extérieur
                if smallest_odds <= 200:  # La plus petite cote est <= 200
                    if first_odds == smallest_odds:  # Favoris domicile (première cote)
                        favoris_domicile_moins_200.append([score, odds, ou_2_5, parsed_date])
                    else:  # Favoris extérieur (dernière cote)
                        favoris_exterieurs_moins_200.append([score, odds, ou_2_5, parsed_date])
                else:  # La plus petite cote > 200
                    if first_odds == smallest_odds:  # Favoris domicile (première cote)
                        favoris_domicile_plus_200.append([score, odds, ou_2_5, parsed_date])
                    else:  # Favoris extérieur (dernière cote)
                        favoris_exterieurs_plus_200.append([score, odds, ou_2_5, parsed_date])

            else:
                print(f"Skipping invalid odds: {odds} (expected 3 values but found {len(odds_values)})")

        except Exception as e:
            print(f"Error processing row {row}: {e}")

# Créer une nouvelle feuille pour chaque catégorie et y insérer les données
def create_sheet_and_add_data(sheet_name, data):
    sheet = wb.create_sheet(sheet_name)
    sheet.append(['Score', '1XBET ODDS', '1XBET O/U 2.5', 'Date'])
    
    # Grouping by 1XBET O/U 2.5 and adding separator lines
    grouped_data = defaultdict(list)
    for row in data:
        ou_2_5 = row[2]
        grouped_data[ou_2_5].append(row)

    # Trier les groupes par "1XBET O/U 2.5" (du plus petit au plus grand)
    sorted_grouped_data = sorted(grouped_data.items(), key=lambda x: x[0])

    # Ajouter les données dans la feuille avec un séparateur
    for ou_2_5, group in sorted_grouped_data:
        for row in group:
            sheet.append(row)
        sheet.append(['--------------------------------------------'])  # Ajouter une ligne de séparation pour chaque groupe


# Créer les onglets avec les résultats
create_sheet_and_add_data('Favoris Domicile Moins de 200', favoris_domicile_moins_200)
create_sheet_and_add_data('Favoris Domicile Plus de 200', favoris_domicile_plus_200)
create_sheet_and_add_data('Favoris Exterieurs Moins de 200', favoris_exterieurs_moins_200)
create_sheet_and_add_data('Favoris Exterieurs Plus de 200', favoris_exterieurs_plus_200)

# Sauvegarder le fichier avec les nouvelles feuilles
wb.save('resultats_favoris_mois.xlsx')

print("Les résultats ont été enregistrés dans 'resultats_favoris_mois.xlsx'")
