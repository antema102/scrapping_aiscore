import openpyxl
from openpyxl import Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import time
import requests
import os

# Définir le chemin vers le ChromeDriver
chrome_driver_path = r"C:\Users\etech\Downloads\Nouveau dossier\chromedriver-win64\chromedriver-win64\chromedriver.exe" 
service = Service(chrome_driver_path)

# Fonction pour vérifier la connexion Internet
def check_internet_connection(url='http://www.google.com', timeout=5):
    try:
        requests.get(url, timeout=timeout)
        return True
    except requests.ConnectionError:
        return False

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique
def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r',encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique
def save_processed_element(element, filename):
    with open(filename, 'a',encoding="utf-8") as file:
        file.write(f"{element}\n")

# Fonction pour faire défiler jusqu'à la fin de la page si plus aucun élément ne se charge
def scroll_to_last_page(driver, max_scroll_attempts=20):
    previous_height = driver.execute_script("return document.body.scrollHeight")
    previous_match_count = 0  # Variable pour suivre le nombre d'éléments 'match-container'
    
    scroll_attempts = 0  # Initialisation d'un compteur pour les tentatives de défilement

    while scroll_attempts < max_scroll_attempts:
        # Obtenez tous les éléments 'match-container' sur la page
        els = driver.execute_script("return document.getElementsByClassName('match-container');")
        current_match_count = len(els)  # Comptez les éléments 'match-container' actuels
        
        if current_match_count > previous_match_count:
            # Faire défiler jusqu'au dernier élément 'match-container'
            driver.execute_script("arguments[0].scrollIntoView();", els[-1])
            time.sleep(2)  # Attendre un peu pour permettre le chargement des nouveaux éléments
            
            # Vérifier la nouvelle hauteur de la page après le défilement
            new_height = driver.execute_script("return document.body.scrollHeight")
            
            # Si la hauteur de la page n'a pas changé, alors on est probablement à la dernière page
            if new_height == previous_height:
                print("Dernière page atteinte ou pas de contenu supplémentaire.")
                break  # Sortir de la boucle car on est à la fin de la page

            # Mettre à jour la hauteur précédente pour le prochain défilement
            previous_height = new_height
            previous_match_count = current_match_count  # Mettre à jour le nombre d'éléments trouvés
        else:
            print("Aucun nouvel élément trouvé. Dernière page atteinte.")
            break  # Aucun nouvel élément trouvé, on arrête la boucle

        scroll_attempts += 1  # Incrémenter le nombre de tentatives de défilement

    # Si on atteint la limite d'essais sans changement détecté
    if scroll_attempts >= max_scroll_attempts:
        print("Limite de défilement atteinte sans changement détecté.")
        
def process_url(url):
    date_str = url.split('/')[-1]
    excel_filename = f"match_data_{date_str}.xlsx"
    processed_filename = f"processed_elements_{date_str}.txt"

    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service)

    if os.path.exists(excel_filename):
        wb = load_workbook(excel_filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Match Data"
        ws.append(["SCORE","BET 365 ODDS", "BET365 O/U 2.5","1XBET ODDS","1XBET O/U 2.5","Nom equipe"])

    processed_elements = load_processed_elements(processed_filename)
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#app")))

        tous_selector = "#app > DIV:nth-of-type(3) > DIV:nth-of-type(2) > DIV:nth-of-type(1) > DIV:nth-of-type(2) > DIV:nth-of-type(2) > DIV:nth-of-type(1) > DIV:nth-of-type(1) > SPAN:nth-of-type(1)"
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, tous_selector))).click()

        second_selector = "#app > DIV:nth-of-type(3) > DIV:nth-of-type(2) > DIV:nth-of-type(1) > DIV:nth-of-type(2) > DIV:nth-of-type(2) > DIV:nth-of-type(1) > DIV:nth-of-type(2) > DIV:nth-of-type(2) > LABEL > SPAN > SPAN"
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, second_selector))).click()

        total_matches = 0  # Compteur pour les matchs à traiter

        while True:
            while not check_internet_connection():
                print("Pas de connexion Internet. Tentative de reconnexion dans 5 secondes...")
                time.sleep(5)
                
            match_elements = driver.find_elements(By.CSS_SELECTOR, 'a.match-container')
            new_data_found = False

            # Traiter les éléments de match
            for element in match_elements:
                nomEquipe = element.find_element(By.CSS_SELECTOR, 'span.name.minitext.maxWidth160').text.strip()

                if nomEquipe in processed_elements:
                    print(f"L'équipe {nomEquipe} a déjà été traitée, passage au suivant.")
                    continue  # Passe au prochain élément sans traiter

                href = element.get_attribute("href")
                processed_elements.add(nomEquipe)
                save_processed_element(nomEquipe, processed_filename)
                new_data_found = True

                # Ouvre le lien et récupère les données
                driver.execute_script("window.open(arguments[0]);", href)
                driver.switch_to.window(driver.window_handles[-1])

                # Cliquer sur la troisieme bouton
                thirst_selector = "#app > div.detail.view.border-box.back > div.tab-bar > div > div > a:nth-child(2)"
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, thirst_selector))).click()

                
                #affiche les cotes lorsque je clique sur le bouton cote
                score1 = '#app > div.detail.view.border-box.back > div.top.color-333.flex-col.flex.align-center > div.flex.w-bar-100.homeBox > div.h-top-center.matchStatus3 > div.font-bold.home-score > span'
                score2 = '#app > div.detail.view.border-box.back > div.top.color-333.flex-col.flex.align-center > div.flex.w-bar-100.homeBox > div.h-top-center.matchStatus3 > div.font-bold.away-score > span'
                cote_selectorBet365_1 = '#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(1) > div.flex-1 > div > div:nth-child(1) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(1) > span > span'
                cote_selectorBet365_2 = '#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(1) > div.flex-1 > div > div:nth-child(1) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(2) > span'
                cote_selectorBet265_3 = '#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(1) > div.flex-1 > div > div:nth-child(1) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(3) > span > span'
                
                cote_nombreButBet365="#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(1) > div.flex-1 > div > div:nth-child(3) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(1) > span"
                cote_plusBut365="#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(1) > div.flex-1 > div > div:nth-child(3) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(2) > span"
                cote_moinsBut365="#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(1) > div.flex-1 > div > div:nth-child(3) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(3) > span"

                cote_selector1xBet_1= '#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(2) > div.flex-1 > div > div:nth-child(1) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(1) > span > span'
                cote_selector1xBet_2 ='#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(2) > div.flex-1 > div > div:nth-child(1) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(2) > span'
                cote_selector1xBet_3 ='#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(2) > div.flex-1 > div > div:nth-child(1) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(3) > span > span'

                cote_nombreBut1xBet='#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(2) > div.flex-1 > div > div:nth-child(3) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(1) > span'
                cote_plusBut1xBet='#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(2) > div.flex-1 > div > div:nth-child(3) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(2) > span'
                cote_moinBut1xBet='#app > div.detail.view.border-box.back > div.content-box > span > div > div.newOdds > div:nth-child(3) > div:nth-child(2) > div.flex-1 > div > div:nth-child(3) > div.box.flex.w100.brr.preMatchBg1 > div > div:nth-child(3) > span'
                try:
                    score_text = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, score1))).text.strip()
                    score_text1 = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, score2))).text.strip()

                    try:
                        cote_Bet365 = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_selectorBet365_1))).text.strip().replace('.', '')
                        cote_Bet365_2 = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_selectorBet365_2))).text.strip().replace('.', '')
                        cote_Bet365_3 = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_selectorBet265_3))).text.strip().replace('.', '')
                        cotesBet365 = f"{cote_Bet365}/{cote_Bet365_2}/{cote_Bet365_3}"
                    except Exception:
                        cotesBet365 = ''

                    try:
                        # Récupère la donnée brute de nombre de buts
                        nombreButBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_nombreButBet365))).text.strip()

                        # Initialisation des valeurs
                        TotalButBet365 = ''

                        # Vérifie s'il y a un '/' pour extraire les deux nombres
                        if '/' in nombreButBet:
                            # Sépare les deux valeurs
                            valeurs = nombreButBet.split('/')
                            premier_nombre = float(valeurs[0])
                            second_nombre = float(valeurs[1])

                            # Vérifie si les deux valeurs sont entre 2 et 3
                            if 2 <= premier_nombre <= 3 and 2 <= second_nombre <= 3:
                                # Récupère les valeurs pour plus et moins si les conditions sont respectées
                                plusButBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_plusBut365))).text.strip().replace('.', '')
                                moinsButBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_moinsBut365))).text.strip().replace('.', '')
                                TotalButBet365 = f"{plusButBet}/{moinsButBet}"
                        else:
                            # Si aucun '/' alors on vérifie si le nombre unique est entre 2 et 3
                            unique_nombre = float(nombreButBet)
                            if 2 <= unique_nombre <= 3:
                                plusButBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_plusBut365))).text.strip().replace('.', '')
                                moinsButBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_moinsBut365))).text.strip().replace('.', '')
                                TotalButBet365 = f"{plusButBet}/{moinsButBet}"

                    except Exception:
                        TotalButBet365 = ''  # En cas d'erreur, TotalButBet365 est vide
                    
                    try:
                        cote_1xBet= WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_selector1xBet_1))).text.strip().replace('.', '')
                        cote_1xBet_2= WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_selector1xBet_2))).text.strip().replace('.', '')
                        cote_1xBet_3= WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_selector1xBet_3))).text.strip().replace('.', '')
                        cotes1xBet = f"{cote_1xBet}/{cote_1xBet_2}/{cote_1xBet_3}"  
                    except Exception:
                        cotes1xBet = ''
                    
                    try:
                        # Récupère la donnée brute de nombre de buts
                        nombreBut1xBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_nombreBut1xBet))).text.strip()

                        # Initialisation de la variable
                        TotalBut1xBet = ''

                        # Vérifie s'il y a un '/' pour extraire les deux nombres
                        if '/' in nombreBut1xBet:
                            # Sépare les deux valeurs
                            valeurs = nombreBut1xBet.split('/')
                            premier_nombre = float(valeurs[0])
                            second_nombre = float(valeurs[1])

                            # Vérifie si les deux valeurs sont entre 2 et 3
                            if 2 <= premier_nombre <= 3 and 2 <= second_nombre <= 3:
                                # Récupère les valeurs pour plus et moins si les conditions sont respectées
                                plusBut1xBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_plusBut1xBet))).text.strip().replace('.', '')
                                moinsBut1xBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_moinBut1xBet))).text.strip().replace('.', '')
                                TotalBut1xBet = f"{plusBut1xBet}/{moinsBut1xBet}"
                        else:
                            # Si aucun '/' alors on vérifie si le nombre unique est entre 2 et 3
                            unique_nombre = float(nombreBut1xBet)
                            if 2 <= unique_nombre <= 3:
                                plusBut1xBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_plusBut1xBet))).text.strip().replace('.', '')
                                moinsBut1xBet = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, cote_moinBut1xBet))).text.strip().replace('.', '')
                                TotalBut1xBet = f"{plusBut1xBet}/{moinsBut1xBet}"

                    except Exception:
                        TotalBut1xBet = ''  # En cas d'erreur, TotalBut1xBet est vide

                    ws.append([f"{score_text}-{score_text1}",cotesBet365,TotalButBet365,cotes1xBet,TotalBut1xBet,nomEquipe])
                    wb.save(excel_filename)
                except Exception as e:
                    print("match reporté")

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

            scroll_to_last_page(driver)
    finally:
        driver.quit()
        

# URLs à traiter
urls = ["https://www.aiscore.com/fr/20241020"]

# Traiter chaque URL et enregistrer les données
for url in urls:
    process_url(url)
