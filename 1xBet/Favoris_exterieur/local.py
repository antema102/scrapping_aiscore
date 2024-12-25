import openpyxl

def filter_odds_and_sort():
    # Ouvrir le fichier Excel et la feuille active
    wb = openpyxl.load_workbook('1xBet.xlsx')  # Remplacez par le chemin de votre fichier Excel
    sheet = wb.active

    # Extraire les données de la feuille
    data = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True))
    
    # Obtenir les valeurs uniques dans la colonne "1XBET O/U 2.5", en ignorant les vides
    unique_odds = sorted(set(str(row[2]) for row in data if row[2] is not None and row[2] != ""))

    # Créer ou réinitialiser les feuilles pour les résultats
    if 'Favoris domicile' in wb.sheetnames:
        home_sheet = wb['Favoris domicile']
        for row in home_sheet.iter_rows(min_row=1, max_row=home_sheet.max_row):
            for cell in row:
                cell.value = None  # Vider les données existantes
    else:
        home_sheet = wb.create_sheet('Favoris domicile')

    if 'Favoris extérieur' in wb.sheetnames:
        away_sheet = wb['Favoris extérieur']
        for row in away_sheet.iter_rows(min_row=1, max_row=away_sheet.max_row):
            for cell in row:
                cell.value = None  # Vider les données existantes
    else:
        away_sheet = wb.create_sheet('Favoris extérieur')

    # Ajouter les en-têtes
    home_sheet.append(["Score", "1XBET ODDS", "1XBET O/U 2.5", "Date"])
    away_sheet.append(["Score", "1XBET ODDS", "1XBET O/U 2.5", "Date"])

    # Parcourir chaque valeur unique dans "1XBET O/U 2.5" triée
    for odd in unique_odds:
        # Filtrer les lignes correspondantes
        matching_rows = [row for row in data if str(row[2]) == odd]

        # Trier les lignes en fonction des valeurs de "1XBET ODDS" (du plus petit au plus grand)
        sorted_rows = sorted(
            matching_rows,
            key=lambda row: [int(i) for i in (row[1].split('/') if row[1] else []) if i.isdigit()]  # Vérification de None
        )

        # Ajouter une ligne vide avant chaque groupe de valeurs
        home_sheet.append(["-----------", "--------", "-------", "-------"])
        away_sheet.append(["-----------", "--------", "-------", "-------"])

        # Trier les lignes pour Favoris extérieur et domicile
        home_rows = []
        away_rows = []

        for row in sorted_rows:
            if row[1]:  # Vérifier que "1XBET ODDS" n'est pas vide
                odds = [int(i) for i in row[1].split('/') if i.isdigit()]
                if len(odds) >= 3:  # S'assurer qu'il y a au moins 3 cotes
                    if odds[0] == min(odds):  # Si la première cote est la plus petite
                        home_rows.append(row)
                    elif odds[2] == min(odds):  # Si la troisième cote est la plus petite
                        away_rows.append(row)

        # Trier les lignes des favoris extérieur par le dernier chiffre de 1XBET ODDS
        away_rows_sorted = sorted(
            away_rows,
            key=lambda row: int(row[1].split('/')[-1]) if row[1] and row[1].split('/')[-1].isdigit() else float('inf')
        )

        # Ajouter les lignes triées dans les feuilles respectives
        for row in home_rows:
            home_sheet.append(row)

        for row in away_rows_sorted:
            away_sheet.append(row)

    # Sauvegarder le fichier modifié
    wb.save('file_updated.xlsx')  # Remplacez par le chemin du fichier de sortie

    print("Filtrage, tri et ajout dans les onglets terminé!")

# Exécution de la fonction
filter_odds_and_sort()
