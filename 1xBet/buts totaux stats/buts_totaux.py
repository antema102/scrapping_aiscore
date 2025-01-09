import pandas as pd
import gspread
from gspread_formatting import *
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import io
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Configuration des accès à Google Sheets

# Fonction pour déterminer si c'est domicile ou extérieur
def determine_location(odds):
    if isinstance(odds, str) and "/" in odds:
        parts = list(map(int, odds.split("/")))
        if min(parts) == parts[0]:
            return "domicile"
            
        elif min(parts) == parts[-1]:
            return "extérieur"
    return None

# Fonction pour calculer les stats Under/Over
def calculate_stats(df):
    
    stats = []
    for _, row in df.iterrows():
        if row.iloc[2]:
            if isinstance(row.iloc[0], str) and "-" in row.iloc[0]:
                    score1, score2 = map(int, row[0].split("-"))
                    total_buts = score1 + score2
                    under_condition = 1 if total_buts < 3 else 0
                    over_condition = 1 if total_buts > 2 else 0
                    stats.append({
                        "Cotes": row.iloc[2],
                        "Mois": row["Mois"],
                        "Under": under_condition,
                        "Over": over_condition
                     })
            else:
                print("error")

            
    return pd.DataFrame(stats)

# Fonction pour organiser les données par mois
def format_stats_for_excel(df):
    # Ordre chronologique des mois
    ordre_mois = [
    "September", "October", "November", "December","January"
    ]
    df["Mois"] = pd.Categorical(df["Mois"], categories=ordre_mois, ordered=True)
    stats_by_month = df.pivot_table(index="Cotes", columns="Mois", values=["Under", "Over"], aggfunc="sum", fill_value=0)
    stats_by_month = stats_by_month.reindex(columns=ordre_mois, level=1)

    return stats_by_month

# Fonction pour ajouter un tableau avec un format spécifique
def add_stats_to_sheet(sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)
    
    # Ajouter les titres des mois et fusionner les cellules pour "Under" et "Over"
    months = data.columns.levels[1].tolist()  # Extraire les mois
    col_start = 2
    
    for month in months:
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 1)
        cell = ws.cell(row=1, column=col_start)
        cell.value = month
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les titres Under et Over sous chaque mois
        ws.cell(row=2, column=col_start).value = "UNDER"
        ws.cell(row=2, column=col_start + 1).value = "OVER"
        
        col_start += 2
    
    # Remplir les données de "Cotes" et les stats "Under" et "Over"
    row_num = 3
    for idx, row in data.iterrows():
        ws.cell(row=row_num, column=1).value = idx
        col_num = 2
        for month in months:
            under_val = row[("Under", month)]
            over_val = row[("Over", month)]
            # Remplir les valeurs "Under" et "Over"
            under_cell = ws.cell(row=row_num, column=col_num)
            over_cell = ws.cell(row=row_num, column=col_num + 1)
            
            under_cell.value = under_val
            over_cell.value = over_val 

            # Coloration conditionnelle
            if under_val > over_val:
                under_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif over_val > under_val:
                over_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif under_val== 0 and over_val == 0:
                pass
            else:
                under_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                over_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            col_num += 2
        row_num += 1

# Fonction pour télécharger ou mettre à jour un fichier
def upload_to_drive(file_path, mime_type, file_name, parent_id_folder):
    drive_service = build('drive', 'v3', credentials=credentials)

    # Rechercher un fichier avec le même nom dans le dossier
    query = f"'{parent_id_folder}' in parents and name = '{file_name}'"
    
    try:
        # Chercher des fichiers dans le dossier avec ce nom
        results = drive_service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get('files', [])

        if files:
            # Si un fichier existe déjà avec ce nom, on le met à jour
            file_id = files[0]['id']
            print(f"Le fichier avec le nom '{file_name}' existe déjà, mise à jour en cours...")

            media = MediaFileUpload(file_path, mimetype=mime_type)

            # Mise à jour du fichier
            updated_file = drive_service.files().update(
                fileId=file_id,
                media_body=media,
                fields='id'
            ).execute()

            print(f"Fichier mis à jour avec succès sur Google Drive avec ID : {updated_file['id']}")

        else:
            # Si le fichier n'existe pas, on le crée
            print(f"Aucun fichier trouvé avec le nom '{file_name}', création d'un nouveau fichier...")

            file_metadata = {
                'name': file_name,
                'mimeType': "application/vnd.google-apps.spreadsheet",  # Convertir en Google Sheets
                'parents': [parent_id_folder]
            }
            media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # Création du fichier
            new_file = drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()

            print(f"Fichier téléchargé sur Google Drive avec ID : {new_file['id']}")

    except Exception as error:
        print(f"Une erreur est survenue : {error}")

try:
    scope = ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file"]
    credentials = Credentials.from_service_account_file(r"C:\Users\Administrator\Desktop\scrapping_aiscore\xbet-identifiants.json", scopes=scope)
    gc = gspread.authorize(credentials)

    # Charger le fichier Google Sheets existant
    spreadsheet_id = "1ROzI-Xnz-Y4y-QGP_-KsCir3FBquSCNteUIWt2AB5DM"
    spreadsheet = gc.open_by_key(spreadsheet_id)

    parent_id_floder="1c94YkueTJmw4yGeiXEPNQd8vRb7ecu8A"

    # Charger la première feuille dans un DataFrame
    worksheet = spreadsheet.sheet1 # Remplacez par le nom de la feuille si nécessaire
    data = worksheet.get_all_values()

    # Utiliser la première ligne comme en-têtes
    columns = data[0]
    data = data[1:]
    
    df = pd.DataFrame(data, columns=columns)

    # Convertir la date en format datetime et extraire le mois
    df["Date"] = pd.to_datetime(df["Date"], format="%d/%m/%Y")
    df["Mois"] = df["Date"].dt.month_name()

    # Ajouter une colonne pour domicile/extérieur
    df["Location"] = df["1XBET ODDS"].apply(determine_location)

    #Home favorites
    home_less_150 = df[df["Location"] == "domicile"]
    home_less_150_stats = calculate_stats(home_less_150[home_less_150["1XBET ODDS"].apply(lambda x: isinstance(x, str) and int(x.split("/")[0]) <= 150)])

    home_space_betwen_150_200 = df[df["Location"] == "domicile"]
    home_space_betwen_150_200_stats = calculate_stats(home_space_betwen_150_200[home_space_betwen_150_200["1XBET ODDS"].apply(lambda x: isinstance(x, str) and 150 < int(x.split("/")[0]) <= 200)])

    home_plus_200 = df[df["Location"] == "domicile"]
    home_plus_200_stats = calculate_stats(home_plus_200[home_plus_200["1XBET ODDS"].apply(lambda x: isinstance(x, str) and int(x.split("/")[0]) > 200 )])

    #Away Favoris
    away_greater_150 = df[df["Location"] == "extérieur"]
    away_greater_150_stats = calculate_stats(away_greater_150[away_greater_150["1XBET ODDS"].apply(lambda x: isinstance(x, str) and int(x.split("/")[2]) <= 150)])

    away_greater_betwen_150_200 = df[df["Location"] == "extérieur"]
    away_greater_betwen_150_200_stats = calculate_stats(away_greater_betwen_150_200[away_greater_betwen_150_200["1XBET ODDS"].apply(lambda x: isinstance(x, str) and 150 < int(x.split("/")[2]) <= 200)])

    away_greater_200_plus = df[df["Location"] == "extérieur"]
    away_greater_200_plus_stats = calculate_stats(away_greater_200_plus[away_greater_200_plus["1XBET ODDS"].apply(lambda x: isinstance(x, str) and 150 < int(x.split("/")[2]) > 200)])

    # Formater les stats pour chaque condition (domicile/extérieur)
    home_less_150_stats_formatted = format_stats_for_excel(home_less_150_stats)
    home_space_betwen_150_200_stats_formatted = format_stats_for_excel(home_space_betwen_150_200_stats)
    home_plus_200_stats_formatted = format_stats_for_excel(home_plus_200_stats)

    #Away Favorites
    away_greater_150_stats_formatted = format_stats_for_excel(away_greater_150_stats)
    away_greater_betwen_150_200_stats_formatted = format_stats_for_excel(away_greater_betwen_150_200_stats)
    away_greater_200_plus_stats_formatted = format_stats_for_excel(away_greater_200_plus_stats)

    # Créer un fichier Excel avec openpyxl
    wb = Workbook()

    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # Ajouter les données formatées dans des feuilles différentes
    add_stats_to_sheet("Domicile  x <= 150", home_less_150_stats_formatted)
    add_stats_to_sheet("Domicile 150 < x <= 200", home_space_betwen_150_200_stats_formatted)
    add_stats_to_sheet("Domicile  x > 200", home_plus_200_stats_formatted)

    add_stats_to_sheet("Extérieur x <= 150", away_greater_150_stats_formatted)
    add_stats_to_sheet("Extérieur 150 < x <= 200", away_greater_betwen_150_200_stats_formatted)
    add_stats_to_sheet("Extérieur > 200", away_greater_200_plus_stats_formatted)

    # Sauvegarder le fichier Excel
    wb.save("C:\\Users\\Administrator\\Desktop\\scrapping_aiscore\\1xBet\\buts totaux stats\\cotes_stats_formatte.xlsx")

    # Exemple d'appel de la fonction
    file_name = 'stats_buts_totaux_Domcile_extérieurs'

    upload_to_drive("C:\\Users\\Administrator\\Desktop\\scrapping_aiscore\\1xBet\\buts totaux stats\\cotes_stats_formatte.xlsx", "application/vnd.google-apps.spreadsheet", file_name, parent_id_floder)

    print("Analyse terminée et exportée vers 'cotes_stats_formatte.xlsx'.")

except Exception as e:
    print("error",e)