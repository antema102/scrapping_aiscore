import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill


# Charger le fichier Excel
df = pd.read_excel("../1xBet.xlsx")

# Convertir la date en format datetime et extraire le mois
df["Date"] = pd.to_datetime(df["Date"], format="%d/%m/%Y")
df["Mois"] = df["Date"].dt.month_name()

# Fonction pour déterminer si c'est domicile ou extérieur
def determine_location(odds):
    if isinstance(odds, str) and "/" in odds:
        parts = list(map(int, odds.split("/")))
        if min(parts) == parts[0]:
            return "domicile"
            
        elif min(parts) == parts[-1]:
            return "extérieur"
    return None

# Ajouter une colonne pour domicile/extérieur
df["Location"] = df["1XBET ODDS"].apply(determine_location)

# Fonction pour calculer les stats Under/Over
def calculate_stats(df):
    stats = []
    for _, row in df.iterrows():
        if isinstance(row.iloc[0], str) and "-" in row.iloc[0]:
            score1, score2 = map(int, row[0].split("-"))
            total_buts = score1 + score2
            under_condition = 1 if total_buts < 3 else 0
            over_condition = 1 if total_buts > 2 else 0
            stats.append({
                "Cotes": row.iloc[2],
                "Mois": row["Mois"],
                "Under": under_condition,
                "Over": over_condition
            })
        
    return pd.DataFrame(stats)

# Filtrer d'abord par la localisation, puis appliquer les conditions sur les cotes

#Home favorites
home_less_150 = df[df["Location"] == "domicile"]
home_less_150_stats = calculate_stats(home_less_150[home_less_150["1XBET ODDS"].apply(lambda x: isinstance(x, str) and int(x.split("/")[0]) <= 150)])

home_space_betwen_150_200 = df[df["Location"] == "domicile"]
home_space_betwen_150_200_stats = calculate_stats(home_space_betwen_150_200[home_space_betwen_150_200["1XBET ODDS"].apply(lambda x: isinstance(x, str) and 150 < int(x.split("/")[0]) <= 200)])

home_plus_200 = df[df["Location"] == "domicile"]
home_plus_200_stats = calculate_stats(home_plus_200[home_plus_200["1XBET ODDS"].apply(lambda x: isinstance(x, str) and int(x.split("/")[0]) > 200 )])

#Away Favoris
away_greater_150 = df[df["Location"] == "extérieur"]
away_greater_150_stats = calculate_stats(away_greater_150[away_greater_150["1XBET ODDS"].apply(lambda x: isinstance(x, str) and int(x.split("/")[2]) <= 150)])

away_greater_betwen_150_200 = df[df["Location"] == "extérieur"]
away_greater_betwen_150_200_stats = calculate_stats(away_greater_betwen_150_200[away_greater_betwen_150_200["1XBET ODDS"].apply(lambda x: isinstance(x, str) and 150 < int(x.split("/")[2]) <= 200)])

away_greater_200_plus = df[df["Location"] == "extérieur"]
away_greater_200_plus_stats = calculate_stats(away_greater_200_plus[away_greater_200_plus["1XBET ODDS"].apply(lambda x: isinstance(x, str) and 150 < int(x.split("/")[2]) > 200)])

# Fonction pour organiser les données par mois
def format_stats_for_excel(df):
    # Ordre chronologique des mois
    ordre_mois = [
    "September", "October", "November", "December"
    ]
    df["Mois"] = pd.Categorical(df["Mois"], categories=ordre_mois, ordered=True)
    stats_by_month = df.pivot_table(index="Cotes", columns="Mois", values=["Under", "Over"], aggfunc="sum", fill_value=0)
    stats_by_month = stats_by_month.reindex(columns=ordre_mois, level=1)

    return stats_by_month

# Formater les stats pour chaque condition (domicile/extérieur)
home_less_150_stats_formatted = format_stats_for_excel(home_less_150_stats)
home_space_betwen_150_200_stats_formatted = format_stats_for_excel(home_space_betwen_150_200_stats)
home_plus_200_stats_formatted = format_stats_for_excel(home_plus_200_stats)

#Away Favorites
away_greater_150_stats_formatted = format_stats_for_excel(away_greater_150_stats)
away_greater_betwen_150_200_stats_formatted = format_stats_for_excel(away_greater_betwen_150_200_stats)
away_greater_200_plus_stats_formatted = format_stats_for_excel(away_greater_200_plus_stats)

# Créer un fichier Excel avec openpyxl
wb = Workbook()

# Supprimer la feuille par défaut
wb.remove(wb.active)


# Fonction pour ajouter un tableau avec un format spécifique
def add_stats_to_sheet(sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)
    
    # Ajouter les titres des mois et fusionner les cellules pour "Under" et "Over"
    months = data.columns.levels[1].tolist()  # Extraire les mois
    col_start = 2
    for month in months:
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 1)
        cell = ws.cell(row=1, column=col_start)
        cell.value = month
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les titres Under et Over sous chaque mois
        ws.cell(row=2, column=col_start).value = "UNDER"
        ws.cell(row=2, column=col_start + 1).value = "OVER"
        
        col_start += 2
    
    # Remplir les données de "Cotes" et les stats "Under" et "Over"
    row_num = 3
    for idx, row in data.iterrows():
        ws.cell(row=row_num, column=1).value = idx
        col_num = 2
        for month in months:
            under_val = row[("Under", month)]
            over_val = row[("Over", month)]
            # Remplir les valeurs "Under" et "Over"
            under_cell = ws.cell(row=row_num, column=col_num)
            over_cell = ws.cell(row=row_num, column=col_num + 1)
            
            under_cell.value = under_val
            over_cell.value = over_val 

                     # Coloration conditionnelle
            if under_val > over_val:
                under_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif over_val > under_val:
                over_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif under_val== 0 and over_val == 0:
                pass
            else:
                under_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                over_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            col_num += 2
        row_num += 1

# Ajouter les données formatées dans des feuilles différentes
add_stats_to_sheet("Domicile  x <= 150", home_less_150_stats_formatted)
add_stats_to_sheet("Domicile 150 < x <= 200", home_space_betwen_150_200_stats_formatted)
add_stats_to_sheet("Domicile  x > 200", home_plus_200_stats_formatted)

add_stats_to_sheet("Extérieur x <= 150", away_greater_150_stats_formatted)
add_stats_to_sheet("Extérieur 150 < x <= 200", away_greater_betwen_150_200_stats_formatted)
add_stats_to_sheet("Extérieur > 200", away_greater_200_plus_stats_formatted)

# Sauvegarder le fichier Excel
wb.save("cotes_stats_formatte.xlsx")

print("Analyse terminée et exportée vers 'cotes_stats_formatte.xlsx'.")
