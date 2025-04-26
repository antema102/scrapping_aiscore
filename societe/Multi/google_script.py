import random
from seleniumwire import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
import os
from multiprocessing import Process, Lock
import requests
import time
import pandas as pd
import os
from bs4 import BeautifulSoup
import urllib.parse
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Verrou global
files_and_sheets = []
lock = Lock()

# Récupérer l'utilisateur courant
user_name = os.getlogin()


for dep in range(13,26):  # Départements de 8 à 12
    dep_formatted = str(dep).zfill(2)
    parts = [f"part_{j}" for j in range(1, 2)]  # Générer part_1 à part_6
    files_and_sheets.append(
        (f"C:/Users/{user_name}/Desktop/scrapping_aiscore/societe/Multi/DEPT/DEPT_{dep_formatted}.xlsx", parts)
    )


# Fonction pour charger les éléments déjà traités depuis un fichier spécifique
def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique


def save_processed_element(element_sirene, element_name, filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{element_sirene} {element_name}\n")


def societe(file_path, sheets):
    try:
        with open("user_agents.txt", "r", encoding="utf-8") as f:
            user_agents = [line.strip() for line in f if line.strip()]

        seleniumwire_options = {
            'proxy': {
                "http": "http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000",
                "https": "http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000",
            },
            'exclude_hosts': ['www.google.com', 'google.com'],
            'verify_ssl': False,
        }


        random_user_agent = random.choice(user_agents)
        chrome_options = uc.ChromeOptions()
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument(
            f"--user-agent={random_user_agent}")
        chrome_options.add_argument("--window-size=800,600")
        chrome_options.add_experimental_option(
            "prefs", {"profile.managed_default_content_settings.images": 2})
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        # Désactiver JavaScript via les préférences
        prefs = {
            "profile.managed_default_content_settings.images": 2
        }

        chrome_options.add_experimental_option("prefs", prefs)
        # Configuration des headers avec le User-Agent
        headers = {
            "User-Agent": random_user_agent
        }
        proxy = {
            'http': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
            'https': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
        }

        # Démarrage du navigateur
        driver = uc.Chrome(options=chrome_options,
                           seleniumwire_options=seleniumwire_options)
        processed_text = os.path.splitext(os.path.basename(file_path))[0]
        number = processed_text.split("_")[-1]
        directory = os.path.join(f"DEPT_{number}")
        
        processed_filename = os.path.join(
            directory, f"{processed_text}_{sheets}.txt")

        new_file_path = os.path.join(
            directory, f"{processed_text}_{sheets}.xlsx")

        processed_elements = load_processed_elements(processed_filename)

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        # Charger ou créer le fichier Excel
        if os.path.exists(new_file_path):
            workbook = load_workbook(new_file_path)
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            workbook = load_workbook(file_path)
            print(f"Le fichier {new_file_path} a été créé.")
            # Parcourir toutes les feuilles et supprimer celles qui ne sont pas 'sheets'
            for feuille in workbook.sheetnames:
                if feuille != sheets:  # Si ce n'est pas l'onglet à garder
                    ws = workbook[feuille]
                    workbook.remove(ws)

        worksheet_name = sheets  # Nom de la feuille à garder dans le fichier Excel
        worksheet = workbook[worksheet_name]

        try:

            total_elements = worksheet.max_row  # Total des éléments dans la source de données
            processed_count = 0       # Compteur des éléments déjà traités

            # si ignoer alors code est for i, row in enumerate(ws[1:], start=2):
            for i, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
                name_company = row[1]
                code_postal = row[3]
                commune = row[4]   # Nom entreprise
                # Convertit en chaîne de caractères
                sirene_number = str(row[0])

                # Prend les 4 derniers chiffres
                last_four_digits_sirene = sirene_number[-4:]

                str_comparaison = f'{last_four_digits_sirene} {name_company}'

                if str_comparaison in processed_elements:
                    processed_count += 1
                    continue

                base_url = 'https://www.google.com/search?q='
                query = f'{name_company} {commune} societe.com'
                encoded_query = urllib.parse.quote_plus(query)
                url = base_url + encoded_query
                driver.get(url)

                try:
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, 'span[jscontroller="msmzHf"] a')))

                    elements = driver.find_elements(
                        By.CSS_SELECTOR, 'span[jscontroller="msmzHf"] a')

                    # Attendre que les éléments soient chargés

                    found_match = False

                    for item in elements:
                        # Vérifier si l'élément est cliquable
                        href = item.get_attribute("href")
                        sirene = href.split('-')[-1]
                        siren_last = sirene.split('.')[0]
                        last_four_digits = str(siren_last)[-4:]

                        if last_four_digits == last_four_digits_sirene:

                            found_match = True
              

                            worksheet.cell(
                                    row=i, column=1, value=siren_last)

                            print(
                                    f"Sirène trouvé : noms {name_company} numero {siren_last}   ligne {i}")
                            workbook.save(
                                    new_file_path)

                            break

      
                                # Si aucun match n'a été trouvé après la boucle

                    if not found_match:
                        print(
                            f"Aucun sirene trouvé pour le nom  {name_company} code postal {code_postal} comune {commune}  ligne {i}")

                    # Attendre avant de passer à l'élément suivant
                    processed_elements.add(
                        f"{last_four_digits_sirene} {name_company}")

                    save_processed_element(
                        last_four_digits_sirene, name_company, processed_filename)
                    processed_count += 1

                    time.sleep(random.uniform(5, 10))
                    # Enregistrer l'élément traité

                except Exception as e:
                    print(
                        f"captacha")
                    driver.close()
                    driver.quit()
                    return False  # Retourne False pour signaler une erreur

            # Vérifiez si tous les éléments ont été traités
            if processed_count >= total_elements:
                print("Tous les éléments ont été traités.")
                driver.close()
                driver.quit()
                print("Script arrêté car aucune correspondance n'a été trouvée.")
                return True

        except Exception as e:
            print(f"Erreur lors de l'exécution", e)
            driver.close()
            driver.quit()  # Nettoyer correctement le driver
            return False  # Retourne False pour signaler une erreur

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)

        return False  # Retourne False pour signaler une erreur


def retry_societe(file_path, sheet_name):
    """
    Fonction pour exécuter et relancer le traitement si une erreur se produit.
    """
    while True:  # Boucle infinie jusqu'à ce que le traitement soit terminé avec succès
        try:
            print(
                f"[INFO] Lancement du traitement : {file_path} - {sheet_name}")
            success = societe(file_path, sheet_name)

            if success:
                print(
                    f"[SUCCESS] Traitement terminé : {file_path} - {sheet_name}")
                break  # Succès, on sort de la boucle
            else:
                print(
                    f"[WARNING] Échec, relance dans 10s : {file_path} - {sheet_name}")
                time.sleep(10)
        except Exception as e:
            print(f"[ERROR] Erreur fatale : {e}")
            time.sleep(10)  # Attendre avant de réessayer


def launch_processes():
    """
    Fonction pour lancer les traitements en simultané.
    """
    # Liste pour stocker les processus

    for file_path, sheets in files_and_sheets:
        dep_number = os.path.basename(file_path).split('_')[1].split('.')[0]
        processes = []

        for sheet_name in sheets:
            print(f"Création d'un processus pour {file_path} - {sheet_name}")

            # Créer un processus pour chaque combinaison fichier/feuille
            process = Process(target=retry_societe,
                              args=(file_path, sheet_name))
            processes.append(process)
            process.start()  # Lancer le processus
            time.sleep(20)

        # Attendre que tous les processus soient terminés
        for process in processes:
            process.join()

        # Générer dynamiquement le nom du fichier fusionné
        departments_str = dep_number  # Ici, juste le département en cours
        directory = os.path.join(f"C:\\", "Users", user_name, "Desktop",
                                 "scrapping_aiscore", "societe", "Multi", f"DEPT_{dep_number}")
        print(f"Traitement du département {dep_number} terminé.")
        output_file = os.path.join(
            directory, f"news_dep_{departments_str}.xlsx")
        print(output_file)

        # Une fois tous les processus terminés, fusionner les fichierss
        merge_excel_files(output_file, dep_number, directory)

    print("Tous les départements ont été traités.")


def send_to_google_sheets(excel_file, dep_number):
    """Ajoute les nouvelles données dans Google Sheets sans effacer l'existant."""
    try:
        # Authentification Google Sheets
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            f'C:/Users/{user_name}/Desktop/scrapping_aiscore/credentials.json', scope
        )
        client = gspread.authorize(creds)
        sheet_id = "1JkycUQRhV7kDnrA-wEfukJAUmEYFj_qFAN6JVx9wVto"
        google_sheet = client.open_by_key(sheet_id)

        # Vérifier si l'onglet existe
        try:
            dep_sheet = google_sheet.worksheet(f"dep_{dep_number}")
            print(
                f"L'onglet dep_{dep_number} existe déjà. Mise à jour en cours...")
        except gspread.exceptions.WorksheetNotFound:
            # Si l'onglet n'existe pas, le créer
            dep_sheet = google_sheet.add_worksheet(
                title=f"dep_{dep_number}", rows="1000", cols="20")
            print(f"L'onglet dep_{dep_number} a été créé.")

        # Lire le fichier Excel
        df = pd.read_excel(excel_file)

        # Remplacer les valeurs NaN par une chaîne vide (évite l'erreur JSON)
        df = df.fillna("")

        # Convertir les données en liste de listes
        data = df.values.tolist()

        # Ajouter les nouvelles données sous l'ancienne
        dep_sheet.append_rows(data, value_input_option="RAW")
        print(
            f"Les données ont été ajoutées avec succès dans l'onglet dep_{dep_number}.")

    except Exception as e:
        print(f"Erreur lors de l'envoi des données à Google Sheets : {e}")

# Fonction pour fusionner les fichiers Excel


def merge_excel_files(output_file, dep_number, directory):
    try:
        all_data = []
        # Parcourir les fichiers générés
        for file_path, parts in files_and_sheets:
            for part_name in parts:
                basename = f"DEPT_{dep_number}_{part_name}"
                individual_file = os.path.join(directory, f"{basename}.xlsx")
                if os.path.exists(individual_file):
                    df = pd.read_excel(individual_file)
                    all_data.append(df)

        # Fusionner toutes les données
        if all_data:
            merged_df = pd.concat(all_data, ignore_index=True)

            print(f"Suppresion des doublons : {output_file}")

            merged_df.drop_duplicates()

            merged_df.to_excel(output_file, index=False)

            print(f"Fichier fusionné créé : {output_file}")
        else:
            print("Aucun fichier à fusionner.")
    except Exception as e:
        print("error merge ")


if __name__ == "__main__":
    print("Lancement des traitements en simultané...")
    launch_processes()
    print("Tous les traitements sont terminés.")
