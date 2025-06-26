import random
from seleniumwire import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import html
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import base64
from urllib.parse import urljoin
import os
import time
import os
import urllib.parse
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Récupérer l'utilisateur courant
user_name = os.getlogin()

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique
def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique
def save_processed_element(element, filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{element}\n")

def human_scroll(driver, steps=10, min_pause=0.2, max_pause=0.7, min_scroll=80, max_scroll=150):
    for _ in range(steps):
        scroll_amount = random.randint(min_scroll, max_scroll)
        driver.execute_script(f"window.scrollBy(0, {scroll_amount});")
        time.sleep(random.uniform(min_pause, max_pause))

def interceptor(request):
    # Bloquer tout sauf le HTML principal
    if request.path.endswith(('.css', '.js', '.png', '.jpg', '.jpeg', '.gif', '.webp', '.woff', '.woff2', '.ttf', '.svg', '.ico')):
        request.abort()
        
def process_url(urls,dep):
    try:
        with open("user_agents.txt", "r", encoding="utf-8") as f:
            user_agents = [line.strip() for line in f if line.strip()]

        
        seleniumwire_options = {
            'proxy': {
                "http": "http://brd-customer-hl_1d81edaa-zone-datacenter_proxy1:wur28vaq23lx@brd.superproxy.io:33335",
                "https": "http://brd-customer-hl_1d81edaa-zone-datacenter_proxy1:wur28vaq23lx@brd.superproxy.io:33335",
            },
            'disable_encoding': True,
        }

        options = {
        'disable_encoding': True,
        }


        random_user_agent = random.choice(user_agents)
        chrome_options = uc.ChromeOptions()
        chrome_options.add_argument("--disable-infobars")
        chrome_options.headless = True
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-modal-animations")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument('--blink-settings=imagesEnabled=false')
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-crash-reporter")
        chrome_options.add_argument("--disable-crashpad-for-testing")
        chrome_options.add_argument("--disable-features=ServiceWorker")
        chrome_options.add_argument(f"--user-agent={random_user_agent}")

        # Désactiver JavaScript via les préférences
        prefs = {
            "profile.managed_default_content_settings.images": 2,
            "profile.managed_default_content_settings.stylesheets": 2,
            "profile.managed_default_content_settings.javascript": 2,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        #Akshay
        # driver_path =f"C:/Users/{user_name}/OneDrive/Desktop/scrapping_aiscore/chromedriver.exe"
        #Local
        driver_path =f"C:/Users/{user_name}/Desktop/scrapping_aiscore/chromedriver.exe"

        driver = uc.Chrome(
            options=chrome_options,
            driver_executable_path=driver_path,
            seleniumwire_options=options
        )

        departement = f"{dep}"
        text_urls = urls.replace(" ", "_")
        directory = os.path.join(f"{text_urls}_{departement}")

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        processed_text = f"processed_elements_{text_urls}.txt"
        excel_filename = f"{text_urls}.xlsx"
        new_file_path = os.path.join(directory, excel_filename)
        processed_filename = os.path.join(directory, processed_text)
        processed_elements = load_processed_elements(processed_filename)

        if os.path.exists(new_file_path):
            wb = load_workbook(new_file_path)
            ws = wb.active
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = text_urls
            ws.append(["Sirene","Dénomination", "Liens Pages Jaunes","Numero","Addresse","Site Web","Tranche effectif","Forme Juridiques","Date Création","Autre Dénominations"])
            print(f"Fichier Excel créé : {new_file_path}")
        try:
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            driver.request_interceptor = interceptor
            driver.get('https://www.google.fr')
            human_scroll(driver)
            base_url = "https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui="
            encoded_quoiqui = urllib.parse.quote_plus(urls)
            encoded_ou = urllib.parse.quote_plus(departement)
            url = f"{base_url}{encoded_quoiqui}&ou={encoded_ou}"
            driver.get(url)

            while True:
                try:
                    WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located(
                            (By.CSS_SELECTOR, "a.bi-denomination"))
                    )

                    urls_pages_jaunes = driver.find_elements(
                        By.CSS_SELECTOR, "a.bi-denomination")
                    
                    if not urls_pages_jaunes:
                        print("Aucun résultat trouvé.")
                        break

                    print(f"{len(urls_pages_jaunes)} résultats trouvés sur cette page.")

                    for element in urls_pages_jaunes:
                        try:
                            nom = element.text.strip()
                            if not nom:
                                continue

                            if nom in processed_elements:
                                print(f"Nom déjà traité : {nom}")
                                continue

                            print(f"Traitement de : {nom}")
                            save_processed_element(nom, processed_filename)
                            processed_elements.add(nom)

                            href = element.get_attribute("href")
                            data_pjlb = element.get_attribute("data-pjlb")

                            if  href and data_pjlb:
                                try:
                                        decoded_pjlb_str = html.unescape(data_pjlb)
                                        data_pjlb_json = json.loads(decoded_pjlb_str)
                                        encoded_url = data_pjlb_json.get("url")
                                        if encoded_url:
                                            decoded_url = base64.b64decode(encoded_url).decode("utf-8")
                                            full_url = urljoin("https://www.pagesjaunes.fr", decoded_url)
                                            original_tabs = driver.window_handles
                                            driver.execute_script("window.open(arguments[0], '_blank');", full_url)
                                            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(original_tabs))
                                            new_tab = [tab for tab in driver.window_handles if tab not in original_tabs][0]
                                            driver.switch_to.window(new_tab)
                                        else:
                                            print("⚠️ Le champ 'url' est manquant dans data-pjlb.")

                                except Exception as e:
                                        print(f"❌ Erreur lors du décodage de data-pjlb : {e}")
 
                            else:
                                original_tabs = driver.window_handles
                                driver.execute_script("window.open(arguments[0], '_blank');", href)
                                WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > len(original_tabs))
                                new_tab = [tab for tab in driver.window_handles if tab not in original_tabs][0]
                                driver.switch_to.window(new_tab)

                            human_scroll(driver)
                            WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.TAG_NAME, "body"))
                            )

                            try:
                                numero_element = driver.find_element(By.CSS_SELECTOR, 'span.coord-numero-mobile > a')
                                numero_=numero_element.get_attribute('title').split('Appeler')
                                numero=numero_[-1]
                            except:
                                numero = ""

                            try:
                                addresse = driver.find_element(By.CSS_SELECTOR, '.address-container > a > span.noTrad').text
                            except:
                                addresse = ""

                            try:
                                site_url = driver.find_element(By.CSS_SELECTOR, '.lvs-container > a > span.value').text
                            except:
                                site_url = ""

                            siren = forme = date_creation = autres_denom =  tranche_effectif = ""

                            try:
                                dl_element = driver.find_element(By.CSS_SELECTOR, 'dl.info-entreprise')
                                dts = dl_element.find_elements(By.TAG_NAME, 'dt')
                                dds = dl_element.find_elements(By.TAG_NAME, 'dd')
                                nb_paires = min(len(dts), len(dds))
                                for i in range(nb_paires):
                                    dt = dts[i]
                                    dd = dds[i]
                                    label = driver.execute_script("return arguments[0].textContent;", dt).strip()
                                    strongs = dd.find_elements(By.TAG_NAME, 'strong')
                                    if len(strongs) > 1:
                                        value = [driver.execute_script("return arguments[0].textContent;", s).strip() for s in strongs]
                                    elif len(strongs) == 1:
                                        value = driver.execute_script("return arguments[0].textContent;", strongs[0]).strip()
                                    else:
                                        value = driver.execute_script("return arguments[0].textContent;", dd).strip()
                                    if "SIREN" in label:
                                        siren = value
                                    elif "Forme juridique" in label:
                                        forme = value
                                    elif "Création d'entreprise" in label:
                                        date_creation = value
                                    elif "Effectif de l'entreprise" in label:
                                        tranche_effectif = value
                                    elif "Autres dénominations" in label:
                                        autres_denom = value

                            except Exception as e:
                                print(f"pas de donnés INSEE")

                            if isinstance(autres_denom, list):
                                autres_denom_str = ", ".join(autres_denom)
                            else:
                                autres_denom_str = autres_denom if autres_denom else ""

                            current_url = driver.current_url
                            ws.append([siren,nom, current_url, numero, addresse, site_url,tranche_effectif,forme, date_creation, autres_denom_str])
                            wb.save(new_file_path)
                            driver.requests.clear()
                            driver.close()
                            driver.switch_to.window(original_tabs[0])
                            time.sleep(random.uniform(1, 10))

                        except Exception as e:
                            print(f"Erreur sur élément : {e}")
                            if len(driver.window_handles) > 1:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])

                    try:
                        next_button = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.ID, 'pagination-next'))
                        )

                        if next_button:
                            data_pjlb_urls=next_button.get_attribute("data-pjlb")
                            decoded_pjlb_str_urls = html.unescape(data_pjlb_urls)
                            data_pjlb_json_urls = json.loads(decoded_pjlb_str_urls)
                            encoded_url_urls = data_pjlb_json_urls.get("url")
                            if encoded_url_urls:
                                print("➡️ Page suivante...")
                                decoded_url_urls = base64.b64decode(encoded_url_urls).decode("utf-8")
                                full_url_urls = urljoin("https://www.pagesjaunes.fr", decoded_url_urls)
                                driver.requests.clear()
                                driver.get(full_url_urls)
                            else:
                                print("⚠️ Le champ 'url' est manquant dans data-pjlb.")

                        else:
                            print("❌ Fin de pagination.")
                    
                    except Exception as e:
                        print(f"Erreur pagination ou fin des pages")
                        driver.requests.clear()
                        driver.quit()
                        return True
                    
                except Exception as e:
                    print(f"Erreur lors du chargement de la page ,il y a une le cloudflary")
                    break
        
            driver.quit()

        except Exception as e:
            print(f"Erreur lors de l'exécution : {e}")
            driver.quit()
            return False

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)
        driver.quit()
        return False

# Lire les URLs depuis le fichier
with open('categorie.txt', 'r', encoding='utf-8') as file:
    urls = [line.strip() for line in file if line.strip()]

# Charger les URLs déjà traitées
done_file = 'done.txt'
if os.path.exists(done_file):
    with open(done_file, 'r', encoding='utf-8') as f:
        done_urls = set(line.strip() for line in f if line.strip())
else:
    done_urls = set()

# Boucle principale
for url in urls:
    if url in done_urls:
        print(f"[SKIP] {url} déjà traité")
        continue

    print(f"[START] Traitement de : {url}")
    while True:
        result = process_url(url,dep=1)
        if result:
            print(f"[SUCCESS] Fini : {url}")
            # On enregistre dans done.txt
            with open(done_file, 'a', encoding='utf-8') as f:
                f.write(url + '\n')
            break
        else:
            print(f"[RETRY] Échec sur : {url} ... Nouvelle tentative dans 2 minutes.")
            time.sleep(120)
