from openpyxl import load_workbook
import pandas as pd
import os
from multiprocessing import Process, Lock
import time
import pandas as pd
import os
import random
import requests
from bs4 import BeautifulSoup
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Verrou global
files_and_sheets = []
lock = Lock()

# Récupérer l'utilisateur courant
user_name = os.getlogin()

for dep in range(1, 13):  # Départements de 8 à 12
    dep_formatted = str(dep).zfill(2)
    parts = [f"part_{j}" for j in range(1, 11)]  # Générer part_1 à part_6
    files_and_sheets.append(
        (f"C:/Users/{user_name}/Desktop/scrapping_aiscore/societe/Multi/DEPT/DEPT_{dep_formatted}.xlsx", parts)
    )

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique
def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique
def save_processed_element(sirene, filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{sirene}\n")


def societe(file_path, sheets):
    try:
        # Proxy avec authentification (nom d'utilisateur et mot de passe)
        proxy = {
            'http': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
            'https': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
        }

        with open("user_agents.txt", "r", encoding="utf-8") as f:
            user_agents = [line.strip() for line in f if line.strip()]

        user_agent = random.choice(user_agents)

        # Configuration des headers avec le User-Agent
        headers = {
            "User-Agent": user_agent
        }

        processed_text = os.path.splitext(os.path.basename(file_path))[0]
        number = processed_text.split("_")[-1]
        directory = os.path.join(f"DEPT_{number}")
        processed_filename = os.path.join(
            directory, f"{processed_text}_{sheets}.txt")
        new_file_path = os.path.join(
            directory, f"{processed_text}_{sheets}.xlsx")

        processed_elements = load_processed_elements(processed_filename)

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        if os.path.exists(new_file_path):
            workbook = load_workbook(new_file_path,)
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            workbook = load_workbook(file_path)
            print(f"Le fichier {new_file_path} a été créé.")
            for feuille in workbook.sheetnames:
                if feuille != sheets:  
                    ws = workbook[feuille]
                    workbook.remove(ws)

        worksheet_name = sheets 
        worksheet = workbook[worksheet_name]

        urllib3.disable_warnings(
            urllib3.exceptions.InsecureRequestWarning)

        try:
            total_elements = worksheet.max_row  # Total des éléments dans la source de données
            processed_count = 0       # Compteur des éléments déjà traités
            for i, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
                sirene_number = str(row[0]).strip()  

                if sirene_number in processed_elements:
                    processed_count += 1
                    continue

                url = f'https://www.societe.com/cgi-bin/search?q={sirene_number}'
                found_match = False

                try:
                    lien_societe = url
                    response = requests.get(lien_societe, headers=headers, timeout=5, verify=False,proxies=proxy)

                    if response.status_code == 200:

                        soup = BeautifulSoup(
                            response.text, 'html.parser')
                        status_entreprise = soup.select_one(
                            '#identite > header > div > span')
                        
                        if status_entreprise:
                            
                            found_match = True
                            li = soup.select('.co-resume > ul > li')
                            etablissement_texte = soup.select_one('#ariaEstablishmentsH > h3')
                            
                            # Valeurs par défaut
                            adresse_principale = code_postal = libelle_commune = pays = status_entreprises =denominations= "N/A"
                            secteur_activite = span_forme_juridique = tranche_effectif = span_diregeants = code_ape = anne = montant = "N/A"
                            nombres_etablissements = 0

                            denominations = soup.select_one('h1').text.strip()
                            status_entreprises = status_entreprise.text.strip()

                            if etablissement_texte:
                                partie_gauche = etablissement_texte.text.strip().split(',')[0]
                                mots = partie_gauche.split()
                                if mots and mots[0].isdigit():
                                    nombres_etablissements = int(mots[0])

                            for item in li:
                                try:
                                    span_label = item.select_one(
                                        'span.ui-label')
                                    
                                    if not span_label:
                                        continue

                                    span_text = span_label.text.strip()

                                    if span_text == 'ADRESSE':
                                        span_adresse_tag = item.select_one(
                                            'span:nth-child(2) > a')
                                        if span_adresse_tag:
                                            span_adresse = span_adresse_tag.text.strip()
                                            if ',' in span_adresse:
                                                adresse_parts = span_adresse.split(
                                                    ',')
                                                adresse_principale = adresse_parts[0].strip(
                                                )
                                                reste = adresse_parts[1].strip(
                                                )
                                                reste_split = reste.split()

                                                if len(reste_split) >= 2:
                                                    code_postal = reste_split[0]
                                                    mots_apres_cp = reste_split[1:]

                                                    if len(mots_apres_cp) > 1:
                                                        libelle_commune = ' '.join(
                                                            mots_apres_cp[:-1])
                                                        pays = mots_apres_cp[-1]
                                                    else:
                                                        libelle_commune = mots_apres_cp[0]
                                                        pays = ''

                                                    if libelle_commune.upper() == pays.upper():
                                                        libelle_commune = pays
                                            else:
                                                adresse_principale = span_adresse

                                    elif span_text == 'CODE NAF ou APE':
                                        span_code_naf = item.select_one(
                                            'span:nth-child(2)')
                                        if span_code_naf:
                                            parts = span_code_naf.text.strip().split('-')
                                            if len(parts) >= 2:
                                                code_ape = parts[0].strip()
                                                secteur_activite = parts[-1].strip()

                                    elif span_text == 'FORME JURIDIQUE':
                                        span_form = item.select_one(
                                            'span:nth-child(2)')
                                        if span_form:
                                            span_forme_juridique = span_form.text.strip()

                                    elif span_text == 'DIRIGEANTS':
                                        dirigeant_tag = item.select_one(
                                            'span:nth-child(2) > a:nth-child(1)')
                                        if dirigeant_tag:
                                            span_diregeants = dirigeant_tag.text.strip()

                                    elif span_text == 'DATE DE CREATION':
                                        date_tag = item.select_one(
                                            'span:nth-child(2)')
                                        if date_tag:
                                            anne = date_tag.text.strip()

                                except Exception as e:
                                    print("Erreur dans boucle `li` :", e)

                            # Données effectif
                            co = soup.select('co-summary-board li')
                            for item_co in co:
                                try:
                                    div = item_co.select_one('div')
                                    if div and 'Effectif' in div.text:
                                        span = item_co.select_one('span')
                                        if span:
                                            tranche_effectif = span.text.strip()
                                except Exception as e:
                                    print("Erreur effectif :", e)

                            #Chifrres d'affaire
                            chiffre=soup.select('.Table__leader__title.border-white')
                            
                            for item_chiffre in chiffre:
                                try:
                                    label_chiffre = item_chiffre.select_one('td:nth-child(1)')
                                    
                                    if not label_chiffre:
                                        continue

                                    if label_chiffre.text.strip() == "Chiffre d'affaires":             
                                        valeur = item_chiffre.select_one('td:nth-child(3)')

                                        if valeur:
                                            lien = valeur.find('a')
                                            if lien and "publication-bilan" in lien.get('href', ''):
                                                pass
                                            else:
                                                texte = valeur.get_text(strip=True).replace('\xa0', ' ').strip()

                                                parts = texte.split()
                                                if len(parts) >= 2:
                                                    number_part = parts[0]
                                                    unit = " ".join(parts[1:])
                                                else:
                                                    number_part = ''.join(filter(str.isdigit, texte))
                                                    unit = ''.join(filter(str.isalpha, texte))

                                                formatted_number = f"{int(number_part):,}".replace(",", " ")
                                                montant = f"{formatted_number} {unit}"

                                except Exception as e:
                                    print("Erreur lors de la récupération du chiffre d'affaires :", e)

                            # print(f"url: {lien_societe}")
                            # print(
                            #     f"nom: {denominations.text.strip() if denominations else 'N/A'}")
                            # print(
                            #     f"status_entreprise: {status_entreprise.text.strip() if status_entreprise else 'N/A'}")
                            # print(
                            #     f"nombres_etablissements: {nombres_etablissements}")
                            # print(
                            #     f"adresse principale: {adresse_principale}")
                            # print(f"code_postal: {code_postal}")
                            # print(f"libelle_commune: {libelle_commune}")
                            # print(f"pays: {pays}")
                            # print(f"secteur_activite: {secteur_activite}")
                            # print(
                            #     f"forme_juridique: {span_forme_juridique}")
                            # print(f"tranche_effectif: {tranche_effectif}")
                            # print(f"dirigeant: {span_diregeants}")
                            # print(f"code_ape: {code_ape}")
                            # print(f"annee_creation: {anne}")
                            # print(f"chiffre d'affaire,{montant}")

                            worksheet.cell(row=i, column=2, value=denominations)
                            worksheet.cell(row=i, column=3, value=adresse_principale)
                            # worksheet.cell(row=i, column=4, value=code_postal)
                            # worksheet.cell(row=i, column=5, value=libelle_commune)
                            # worksheet.cell(row=i, column=6, value=code_ape)
                            # worksheet.cell(row=i, column=7, value=tranche_effectif)
                            worksheet.cell(row=i, column=8, value=status_entreprises)
                            worksheet.cell(row=i, column=9, value=nombres_etablissements)
                            worksheet.cell(row=i, column=10, value=secteur_activite)
                            worksheet.cell(row=i, column=11, value=anne)
                            worksheet.cell(row=i, column=12, value=span_diregeants)
                            worksheet.cell(row=i, column=13, value=span_forme_juridique)
                            worksheet.cell(row=i, column=14, value=montant)
                            workbook.save(new_file_path)
                            print(f"donnés creé {denominations} {i}")
                    else:
                        print(
                            f"⚠️ Statut {response.status_code} pour {lien_societe} {i}")
                        return

                except Exception as e:
                    print("Captcha", e)
                    return False

                if not found_match:
                    print(f"Aucun sirene correspondand {sirene_number}")

                processed_elements.add(sirene_number)
                save_processed_element(sirene_number, processed_filename)
                processed_count += 1

                time.sleep(random.uniform(1,5))

            # Vérifiez si tous les éléments ont été traités
            if processed_count >= total_elements:
                print("Tous les éléments ont été traités.")
                print("Script arrêté car aucune correspondance n'a été trouvée.")
                return True

        except Exception as e:
            print(f"Erreur lors de l'exécution", e)
            return False  # Retourne False pour signaler une erreur

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)
        return False  # Retourne False pour signaler une erreur


def retry_societe(file_path, sheet_name):
    """
    Fonction pour exécuter et relancer le traitement si une erreur se produit.
    """
    while True:  # Boucle infinie jusqu'à ce que le traitement soit terminé avec succès
        try:
            print(
                f"[INFO] Lancement du traitement : {file_path} - {sheet_name}")
            success = societe(file_path, sheet_name)

            if success:
                print(
                    f"[SUCCESS] Traitement terminé : {file_path} - {sheet_name}")
                break  # Succès, on sort de la boucle
            else:
                print(
                    f"[WARNING] Échec, relance dans 10s : {file_path} - {sheet_name}")
                time.sleep(10)

        except Exception as e:
            print(f"[ERROR] Erreur fatale : {e}")
            time.sleep(10)  # Attendre avant de réessayer


def launch_processes():
    """
    Fonction pour lancer les traitements en simultané.
    """
    # Liste pour stocker les processus

    for file_path, sheets in files_and_sheets:
        dep_number = os.path.basename(file_path).split('_')[1].split('.')[0]
        processes = []

        for sheet_name in sheets:
            print(f"Création d'un processus pour {file_path} - {sheet_name}")

            # Créer un processus pour chaque combinaison fichier/feuille
            process = Process(target=retry_societe,
                              args=(file_path, sheet_name))
            processes.append(process)
            process.start()  # Lancer le processus
            time.sleep(20)

        # Attendre que tous les processus soient terminés
        for process in processes:
            process.join()

        # Générer dynamiquement le nom du fichier fusionné
        departments_str = dep_number  # Ici, juste le département en cours
        directory = os.path.join(f"C:\\", "Users", user_name, "Desktop",
                                 "scrapping_aiscore", "societe", "Multi", f"DEPT_{dep_number}")
        print(f"Traitement du département {dep_number} terminé.")
        output_file = os.path.join(
            directory, f"news_dep_{departments_str}.xlsx")
        print(output_file)

        # Une fois tous les processus terminés, fusionner les fichierss
        merge_excel_files(output_file, dep_number, directory)

    print("Tous les départements ont été traités.")


def merge_excel_files(output_file, dep_number, directory):
    try:
        all_data = []
        # Parcourir les fichiers générés
        for file_path, parts in files_and_sheets:
            for part_name in parts:
                basename = f"DEPT_{dep_number}_{part_name}"
                individual_file = os.path.join(directory, f"{basename}.xlsx")
                if os.path.exists(individual_file):
                    df = pd.read_excel(individual_file)
                    all_data.append(df)

        # Fusionner toutes les données
        if all_data:
            merged_df = pd.concat(all_data, ignore_index=True)
            print(f"Suppresion des doublons : {output_file}")
            merged_df.drop_duplicates()
            merged_df.to_excel(output_file, index=False)
            print(f"Fichier fusionné créé : {output_file}")

        else:
            print("Aucun fichier à fusionner.")
    except Exception as e:
        print("error merge ")


if __name__ == "__main__":
    print("Lancement des traitements en simultané...")
    launch_processes()
    print("Tous les traitements sont terminés.")