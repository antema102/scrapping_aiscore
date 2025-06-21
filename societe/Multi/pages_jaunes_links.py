import random
from seleniumwire import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import time
import os
import urllib.parse
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Récupérer l'utilisateur courant
user_name = os.getlogin()

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique
def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique
def save_processed_element(element, filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{element}\n")


def process_url(urls):
    try:
        with open("user_agents.txt", "r", encoding="utf-8") as f:
            user_agents = [line.strip() for line in f if line.strip()]

        seleniumwire_options = {
            'proxy': {
                "http": "http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000",
                "https": "http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000",
            }
        }
        random_user_agent = random.choice(user_agents)
        chrome_options = uc.ChromeOptions()
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-modal-animations")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument('--blink-settings=imagesEnabled=false')
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-crash-reporter")
        chrome_options.add_argument("--disable-crashpad-for-testing")
        chrome_options.add_argument(f"--user-agent={random_user_agent}")
        driver_path = f"C:/Users/{user_name}/Desktop/scrapping_aiscore/chromedriver.exe"

        driver = uc.Chrome(
            options=chrome_options,
            driver_executable_path=driver_path,
        )

        # Désactiver JavaScript via les préférences
        prefs = {
            "profile.managed_default_content_settings.images": 2,
        }

        chrome_options.add_experimental_option("prefs", prefs)
        # service = Service(ChromeDriverManager().install())
        departement = "01"
        text_urls = urls.replace(" ", "_")
        directory = os.path.join(f"{text_urls}_{departement}")

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        processed_text = f"processed_elements_{text_urls}.txt"
        excel_filename = f"{text_urls}.xlsx"
        new_file_path = os.path.join(
            directory, excel_filename)
        processed_filename = os.path.join(
            directory, processed_text)
        processed_elements = load_processed_elements(processed_filename)

        if os.path.exists(new_file_path):
            wb = load_workbook(new_file_path)
            ws = wb.active
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = text_urls
            ws.append(["Noms", "Urls"])
            print(f"Fichier Excel créé : {new_file_path}")

        try:
            base_url = "https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui="
            encoded_quoiqui = urllib.parse.quote_plus(urls)
            encoded_ou = urllib.parse.quote_plus(departement)
            url = f"{base_url}{encoded_quoiqui}&ou={encoded_ou}"
            driver.get(url)
            new_data_found = False
            while True:
                try:
                    WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located(
                            (By.CSS_SELECTOR, "a.bi-denomination"))
                    )

                    urls_pages_jaunes = driver.find_elements(
                        By.CSS_SELECTOR, "a.bi-denomination")
                    
                    if not urls_pages_jaunes:
                        print("Aucun résultat trouvé.")
                        break

                    print(f"{len(urls_pages_jaunes)} résultats trouvés sur cette page.")
                    for element in urls_pages_jaunes:
                        try:
                            nom = element.text.strip()
                            if not nom:
                                continue

                            if nom in processed_elements:
                                print(f"Nom déjà traité : {nom}")
                                continue

                            print(f"Traitement de : {nom}")
                            save_processed_element(nom, processed_filename)
                            processed_elements.add(nom)
                            new_data_found = True

                            # Ouvrir dans un nouvel onglet
                            ActionChains(driver)\
                                .key_down(Keys.CONTROL)\
                                .click(element)\
                                .key_up(Keys.CONTROL)\
                                .perform()

                            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
                            driver.switch_to.window(driver.window_handles[-1])

                            # Attendre chargement + récupérer URL
                            WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.TAG_NAME, "body"))
                            )

                            current_url = driver.current_url
                            ws.append([nom, current_url])
                            wb.save(new_file_path)

                            # Fermer onglet et retour
                            time.sleep(random.uniform(1, 5))
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])

                        except Exception as e:
                            print(f"Erreur sur élément : {e}")
                            if len(driver.window_handles) > 1:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                    try:
                        next_button = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.ID, 'pagination-next'))
                        )
                        if next_button:
                            print("➡️ Page suivante...")
                            next_button.click()
                        else:
                            print("❌ Fin de pagination.")
                            break

                    except Exception as e:
                        print(f"Erreur pagination ou fin des pages : {e}")
                        break

                except Exception as e:
                    print(f"Erreur lors du chargement de la page : {e}")
                    break

            driver.quit()

        except Exception as e:
            print(f"Erreur lors de l'exécution : {e}")

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)


urls = [
    "Garages automobiles réparation"
]

# Traiter chaque URL et enregistrer les données
for url in urls:
    process_url(url)
