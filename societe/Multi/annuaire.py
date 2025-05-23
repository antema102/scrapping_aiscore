from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
import socket
from urllib3.connection import HTTPConnection
import os
from multiprocessing import Process, Lock
import requests
import time
import pandas as pd
import os
import re


# Verrou global
files_and_sheets = []
lock = Lock()

# Définir manuellement l'ID de la VM (1 à 5)
# VM_ID = 3  # Change cette valeur pour 2, 3, 4, ou 5 selon la VM sur laquelle tu exécutes le script

# Nombre total de VMs
# NB_VMS = 5

# Liste des départements (par exemple, de 1 à 20)
# departements = list(range(8,30))  # Exemple : départements de 1 à 20

# Répartition des départements en fonction de l'ID de la VM
# departements_vm = [departements[i] for i in range(VM_ID-1, len(departements), NB_VMS)]

# Récupérer l'utilisateur courant
user_name = os.getlogin()


for dep in range(1, 2):  # Départements de 8 à 12
    dep_formatted = str(dep).zfill(2)
    parts = [f"part_{j}" for j in range(1,2)]  # Générer part_1 à part_6
    files_and_sheets.append(
        (f"C:/Users/{user_name}/Desktop/scrapping_aiscore/societe/Multi/DEPT/DEPT_{dep_formatted}.xlsx", parts)
    )


def check_internet(url="https://www.google.com", timeout=5):
    """Teste la connexion Internet en envoyant une requête à Google."""
    try:
        response = requests.get(url, timeout=timeout)
        return response.status_code == 200
    except requests.ConnectionError:
        return False

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique


def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique


def save_processed_element(element, filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{element}\n")


def societe(file_path, sheets):
    try:
        HTTPConnection.default_socket_options = (
            HTTPConnection.default_socket_options + [
                (socket.SOL_SOCKET, socket.SO_SNDBUF, 1000000),  # 1MB in byte
                (socket.SOL_SOCKET, socket.SO_RCVBUF, 1000000)
            ])

        # Configuration du proxy avec authentification via URL
        seleniumwire_options = {
            'proxy': {
                'http': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
                'https': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
            }
        }

        chrome_options = Options()
        # Dimensions de la fenêtre
        chrome_options.add_argument("--window-size=800,600")
        # Mode sans interface graphique
        # chrome_options.add_argument("--headless")
        # Désactive les barres d'information
        chrome_options.add_argument("--disable-infobars")
        # Empêche la détection d'automatisation
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
        # Pour résoudre certains problèmes de sécurité
        chrome_options.add_argument("--no-sandbox")
        # Évite la mise en arrière-plan des processus de rendu
        chrome_options.add_argument("--disable-renderer-backgrounding")
        # Empêche le ralentissement des minuteries en arrière-plan
        chrome_options.add_argument("--disable-background-timer-throttling")
        # Évite la mise en arrière-plan des fenêtres occultées
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        # Désactive la détection de phishing côté client
        chrome_options.add_argument("--disable-client-side-phishing-detection")
        # Désactive le rapporteur de crash
        chrome_options.add_argument("--disable-crash-reporter")
        # Désactive l'utilisation du GPU pour la compatibilité
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--silent")  # Réduit les logs inutiles
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_experimental_option(
            "excludeSwitches", ["enable-logging"])
        chrome_options.add_argument("--disable-logging")

        # Désactiver JavaScript via les préférences
        prefs = {
            "profile.managed_default_content_settings.images": 2
        }

        chrome_options.add_experimental_option("prefs", prefs)

        service = Service(ChromeDriverManager().install())

        driver = webdriver.Chrome(
            service=service, options=chrome_options, seleniumwire_options=seleniumwire_options)

        processed_text = os.path.splitext(os.path.basename(file_path))[0]

        number = processed_text.split("_")[-1]

        directory = os.path.join(f"DEPT_{number}")

        processed_filename = os.path.join(
            directory, f"{processed_text}_{sheets}.txt")

        new_file_path = os.path.join(
            directory, f"{processed_text}_{sheets}.xlsx")

        processed_elements = load_processed_elements(processed_filename)

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        # Charger ou créer le fichier Excel
        if os.path.exists(new_file_path):
            workbook = load_workbook(new_file_path)
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            workbook = load_workbook(file_path)
            print(f"Le fichier {new_file_path} a été créé.")
            # Parcourir toutes les feuilles et supprimer celles qui ne sont pas 'sheets'
            for feuille in workbook.sheetnames:
                if feuille != sheets:  # Si ce n'est pas l'onglet à garder
                    ws = workbook[feuille]
                    workbook.remove(ws)

        worksheet_name = sheets  # Nom de la feuille à garder dans le fichier Excel
        worksheet = workbook[worksheet_name]

        try:

            total_elements = worksheet.max_row  # Total des éléments dans la source de données
            processed_count = 0       # Compteur des éléments déjà traités

            # si ignoer alors code est for i, row in enumerate(ws[1:], start=2):
            for i, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):

                sirene_number = str(row[0])

                name_company = row[1]  # Nom entreprise
                # Convertit en chaîne de caractères

                cote_postal = row[3]

                comune = row[4]

                # Prend les 4 derniers chiffres
                last_four_digits_sirene = sirene_number[-4:]

                if name_company in processed_elements:
                    processed_count += 1
                    continue

                cta_url = f'https://annuaire-entreprises.data.gouv.fr/rechercher?terme= {name_company} {cote_postal} {comune}'
                driver.get(cta_url)
                
                try:
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a.result-link')))

                    elements = driver.find_elements(By.CSS_SELECTOR, 'a.result-link')

                    try:
                        for item in elements:
                            sirene = item.get_attribute('data-siren')
                            # Prend les 4 derniers chiffres

                            last_four_digits = str(sirene)[-4:]

                            href = item.get_attribute("href")

                            if last_four_digits == last_four_digits_sirene:

                                driver.get(href)

                                WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.ID, "entreprise")))

                                try:
                                    # Trouver tous les éléments <tr>
                                    
                                    tr_list = driver.find_elements(By.CSS_SELECTOR, "#entreprise > div > div:nth-child(2) > table.styleSimple_two-column-table__KG2Pf > tbody > tr")
                                
                                    name_company_str = siren_news = salarier = span_adresse_str = ""

                                    for tr in tr_list:

                                        label_element = tr.find_element(By.CSS_SELECTOR, "td:first-child").text.strip()

                                        if "Dénomination" in label_element:
                                            print("1",label_element)
                                            name_company_str = tr.find_element(By.CSS_SELECTOR, "td:nth-child(2) > button > span")
                                            print(name_company_str)

                                        # elif "SIREN" in label_element:
                                        #     print("1",label_element)
                                        #     siren_news = tr.find_element(By.CSS_SELECTOR, "td:nth-child(2) > button > span").text.strip()

                                        # elif "Effectif salarié" in label_element:
                                        #     print("1",label_element)
                                        #     salarier = tr.find_element(By.CSS_SELECTOR, "td:nth-child(2) > button > span").text.strip()

                                        # elif "Adresse postale" in label_element:
                                        #     try:
                                        #         print("1",label_element)
                                        #         span_adresse_str = tr.find_element(By.CSS_SELECTOR, "td:last-child").text.strip()
                                        #     except:
                                        #         print('pas de nombre effectif')
                                        #         span_adresse_str=''

                                    print(f"Sirène trouvé : noms {name_company_str} numero {siren_news} adresse {span_adresse_str} salarié {salarier} ligne {i}")
                                        
                                    # Sauvegarder le fichier après chaque ligne
                                    workbook.save(new_file_path)

                                except Exception as e:
                                    print('Erreur dans la récupération des données du siren:', e)

                    except Exception as e:
                        print(f"Erreur lors de annuaire",e)

                except Exception as e:
                    try:
                        driver.find_element(By.CSS_SELECTOR, 'div.no-results')
                        print(f"pas de donné dans le recheche")
                        processed_elements.add(name_company)
                        save_processed_element(
                        name_company, processed_filename)
                        processed_count += 1

                    except Exception as e:
                        print("Error",e)
                        driver.close()
                        driver.quit()
                        return False

            # Vérifiez si tous les éléments ont été traités
            if processed_count >= total_elements:
                print("Tous les éléments ont été traités.")
                driver.close()
                driver.quit()
                print("Script arrêté car aucune correspondance n'a été trouvée.")
                return True

        except Exception as e:
            print(f"Erreur lors de l'exécution", e)
            driver.close()
            driver.quit()  # Nettoyer correctement le driver
            return False  # Retourne False pour signaler une erreur

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)

        return False  # Retourne False pour signaler une erreur


def retry_societe(file_path, sheet_name):
    """
    Fonction pour exécuter et relancer le traitement si une erreur se produit.
    """
    while True:  # Boucle infinie jusqu'à ce que le traitement soit terminé avec succès
        try:
            print(
                f"[INFO] Lancement du traitement : {file_path} - {sheet_name}")
            success = societe(file_path, sheet_name)

            if success:
                print(
                    f"[SUCCESS] Traitement terminé : {file_path} - {sheet_name}")
                break  # Succès, on sort de la boucle
            else:
                print(
                    f"[WARNING] Échec, relance dans 10s : {file_path} - {sheet_name}")
                time.sleep(10)
        except Exception as e:
            print(f"[ERROR] Erreur fatale : {e}")
            time.sleep(10)  # Attendre avant de réessayer


def launch_processes():
    """
    Fonction pour lancer les traitements en simultané.
    """
    # Liste pour stocker les processus

    for file_path, sheets in files_and_sheets:
        dep_number = os.path.basename(file_path).split('_')[1].split('.')[0]
        processes = []

        for sheet_name in sheets:
            print(f"Création d'un processus pour {file_path} - {sheet_name}")

            # Créer un processus pour chaque combinaison fichier/feuille
            process = Process(target=retry_societe,
                              args=(file_path, sheet_name))
            processes.append(process)
            process.start()  # Lancer le processus
            time.sleep(20)

        # Attendre que tous les processus soient terminés
        for process in processes:
            process.join()

        # Générer dynamiquement le nom du fichier fusionné
        departments_str = dep_number  # Ici, juste le département en cours
        directory = os.path.join(f"C:\\", "Users", user_name, "Desktop",
                                 "scrapping_aiscore", "societe", "Multi", f"DEPT_{dep_number}")
        print(f"Traitement du département {dep_number} terminé.")
        output_file = os.path.join(
            directory, f"news_dep_{departments_str}.xlsx")
        print(output_file)

        # Une fois tous les processus terminés, fusionner les fichierss
        merge_excel_files(output_file, dep_number, directory)

    print("Tous les départements ont été traités.")


def send_to_google_sheets(excel_file, dep_number):
    """Ajoute les nouvelles données dans Google Sheets sans effacer l'existant."""
    try:
        # Authentification Google Sheets
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            f'C:/Users/{user_name}/Desktop/scrapping_aiscore/credentials.json', scope
        )
        client = gspread.authorize(creds)
        sheet_id = "1JkycUQRhV7kDnrA-wEfukJAUmEYFj_qFAN6JVx9wVto"
        google_sheet = client.open_by_key(sheet_id)

        # Vérifier si l'onglet existe
        try:
            dep_sheet = google_sheet.worksheet(f"dep_{dep_number}")
            print(
                f"L'onglet dep_{dep_number} existe déjà. Mise à jour en cours...")
        except gspread.exceptions.WorksheetNotFound:
            # Si l'onglet n'existe pas, le créer
            dep_sheet = google_sheet.add_worksheet(
                title=f"dep_{dep_number}", rows="1000", cols="20")
            print(f"L'onglet dep_{dep_number} a été créé.")

        # Lire le fichier Excel
        df = pd.read_excel(excel_file)

        # Remplacer les valeurs NaN par une chaîne vide (évite l'erreur JSON)
        df = df.fillna("")

        # Convertir les données en liste de listes
        data = df.values.tolist()

        # Ajouter les nouvelles données sous l'ancienne
        dep_sheet.append_rows(data, value_input_option="RAW")
        print(
            f"Les données ont été ajoutées avec succès dans l'onglet dep_{dep_number}.")

    except Exception as e:
        print(f"Erreur lors de l'envoi des données à Google Sheets : {e}")

# Fonction pour fusionner les fichiers Excel


def merge_excel_files(output_file, dep_number, directory):
    try:
        all_data = []
        # Parcourir les fichiers générés
        for file_path, parts in files_and_sheets:
            for part_name in parts:
                basename = f"DEPT_{dep_number}_{part_name}"
                individual_file = os.path.join(directory, f"{basename}.xlsx")
                if os.path.exists(individual_file):
                    df = pd.read_excel(individual_file)
                    all_data.append(df)

        # Fusionner toutes les données
        if all_data:
            merged_df = pd.concat(all_data, ignore_index=True)

            print(f"Suppresion des doublons : {output_file}")

            merged_df.drop_duplicates()

            merged_df.to_excel(output_file, index=False)

            print(f"Fichier fusionné créé : {output_file}")
        else:
            print("Aucun fichier à fusionner.")
    except Exception as e:
        print("error merge ")


if __name__ == "__main__":
    print("Lancement des traitements en simultané...")
    launch_processes()
    print("Tous les traitements sont terminés.")
