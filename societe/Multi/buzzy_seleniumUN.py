import urllib.parse
from openpyxl import load_workbook
import pandas as pd
import os
from multiprocessing import Process, Lock
import time
import pandas as pd
import os
import random
from bs4 import BeautifulSoup
from seleniumwire import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Verrou global
files_and_sheets = []
lock = Lock()

# Récupérer l'utilisateur courant
user_name = os.getlogin()

for dep in range(75, 76):  # Départements de 8 à 12
    dep_formatted = str(dep).zfill(2)
    parts = [f"part_{j}" for j in range(1, 5)]  # Générer part_1 à part_6
    files_and_sheets.append(
        (f"C:/Users/{user_name}/Desktop/scrapping_aiscore/societe/Multi/DEPT/DEPT_{dep_formatted}.xlsx", parts)
    )

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique


def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique


def save_processed_element(element_sirene,  filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{element_sirene}\n")


def content(soup, title):
    try:
        # Recherche le div avec le bon titlekey
        div = soup.find("div", attrs={"titlekey": title})
        if not div:
            return None
        if title == 'labels.website':
            link = div.select_one("div:last-child > a")
            if link and link.has_attr('href'):
                return link['href']
            else:
                return None
        elif title == 'labels.email':
            link = div.select_one("div:last-child > a")
            if link and link.has_attr('href'):
                href = link['href']
                if href.startswith("mailto:"):
                    email = href.split(':', 1)[-1]
                    return email
            return None
        else:
            p = div.select_one("div:last-child > a > p")
            if p:
                return p.get_text(strip=True)
            else:
                return None

    except Exception as e:
        print(
            f"❌ Erreur lors de l'extraction du contenu pour titlekey = {title} :", e)
        return None


def societe(file_path, sheets):
    try: 
        with open("user_agents.txt", "r", encoding="utf-8") as f:
            user_agents = [line.strip() for line in f if line.strip()]

        seleniumwire_options = {
            'proxy': {
                "http": "http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000",
                "https": "http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000",
            }
        }

        random_user_agent = random.choice(user_agents)
        chrome_options = uc.ChromeOptions()
        chrome_options.headless = True  # Exécuter en mode sans tête
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument(f"--user-agent={random_user_agent}")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument(
            "--disable-blink-features=AutomationControlled")
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-modal-animations")
        chrome_options.add_argument("--disable-logging")
        chrome_options.add_argument('--blink-settings=imagesEnabled=false')
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-crash-reporter")
        chrome_options.add_argument("--disable-crashpad-for-testing")
        # Désactiver JavaScript via les préférences
        prefs = {
            "profile.managed_default_content_settings.images": 2,
            "profile.managed_default_content_settings.stylesheets": 2,
        }
        driver_path = f"C:/Users/{user_name}/Desktop/scrapping_aiscore/chromedriver.exe"
        chrome_options.add_experimental_option("prefs", prefs)

        driver = uc.Chrome(options=chrome_options,
                            driver_executable_path=driver_path,
                           seleniumwire_options=seleniumwire_options)
        driver.set_window_position(-2000, 0)
        processed_text = os.path.splitext(os.path.basename(file_path))[0]
        number = processed_text.split("_")[-1]
        directory = os.path.join(f"DEPT_{number}")
        processed_filename = os.path.join(
            directory, f"{processed_text}_{sheets}.txt")
        new_file_path = os.path.join(
            directory, f"{processed_text}_{sheets}.xlsx")

        processed_elements = load_processed_elements(processed_filename)

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        # Charger ou créer le fichier Excel
        if os.path.exists(new_file_path):
            workbook = load_workbook(new_file_path,)
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            workbook = load_workbook(file_path)
            print(f"Le fichier {new_file_path} a été créé.")
            for feuille in workbook.sheetnames:
                if feuille != sheets:
                    ws = workbook[feuille]
                    workbook.remove(ws)
        worksheet_name = sheets
        worksheet = workbook[worksheet_name]
        try:
            total_elements = worksheet.max_row
            processed_count = 0

            for i, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
                sirene_number = str(row[0])

                if sirene_number in processed_elements:
                    processed_count += 1
                    continue

                url = f'https://bizzy.org/fr/fr/{sirene_number}'
                driver.get(url)
                found_match = False

                try:

                    try:        
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located(
                                (By.CSS_SELECTOR, ".styles_tile__F0oXN"))
                        )

                        html = driver.page_source
                        soup = BeautifulSoup(html, 'html.parser')
                        twitter = youtube = instagram = linkedin = facebook = phone =web = email =None
                        try:
                            phone = content(soup, 'labels.phone-number')
                            web = content(soup, 'labels.website')
                            email = content(soup, 'labels.email')
                            try:
                                reseau = soup.find(
                                    "div", attrs={"titlekey": 'labels.socials'})
                                if not reseau:
                                    pass
                                links = reseau.select("div:last-child a")
                                for a in links:
                                    aria_label = a.get(
                                        "aria-label", "").upper()
                                    if "TWITTER" in aria_label:
                                        twitter = a.get("href")
                                    elif "YOUTUBE" in aria_label:
                                        youtube = a.get("href")
                                    elif "INSTAGRAM" in aria_label:
                                        instagram = a.get("href")
                                    elif "LINKEDIN" in aria_label:
                                        linkedin = a.get("href")
                                    elif "FACEBOOK" in aria_label:
                                        facebook = a.get("href")
                            except:
                                pass

                            if any([phone, web, facebook, twitter, youtube, instagram, linkedin, email]):
                                found_match = True
                                # Affichage final
                                print("phone",phone,"web",web,"facebook",facebook,"twitter",twitter,"youtube",youtube,"instagram",instagram,"linkedin",linkedin,"email",email)

                                worksheet.cell(
                                    row=i, column=8, value=phone)
                                worksheet.cell(row=i, column=9, value=web)
                                worksheet.cell(
                                    row=i, column=10, value=facebook)
                                worksheet.cell(
                                    row=i, column=11, value=twitter)
                                worksheet.cell(
                                    row=i, column=12, value=youtube)
                                worksheet.cell(
                                    row=i, column=13, value=instagram)
                                worksheet.cell(
                                    row=i, column=14, value=linkedin)
                                worksheet.cell(
                                    row=i, column=15, value=email)
                                workbook.save(new_file_path)
                                print(f"donnés creé {sirene_number} {i}")

                        except Exception as e:
                            print(
                                f"Erreur lors de la récupération du SIRENE: {e}")
                    except:
                        try:
                            WebDriverWait(driver, 5).until(
                                EC.presence_of_element_located(
                                    (By.CSS_SELECTOR, "#hs-web-interactives-top-push-anchor"))
                            )
                            print(f"page not found")
                        except :
                            print(f"Captcha coudlfaire")
                            driver.close()
                            driver.quit()
                            return False
                            

                    processed_elements.add(sirene_number)
                    save_processed_element(
                        sirene_number, processed_filename)
                    processed_count += 1
                    time.sleep(random.uniform(1, 5))

                    # Si aucun match n'a été trouvé après la boucle
                    if not found_match:
                            print(
                                f"Aucun Donné trouvé pour le nom  {sirene_number} code postal ligne {i}")

                except Exception as e:
                    print("Captcha", e)
                    driver.close()
                    driver.quit()
                    return False

            # Vérifiez si tous les éléments ont été traités
            if processed_count >= total_elements:
                print("Tous les éléments ont été traités.")
                print("Script arrêté car aucune correspondance n'a été trouvée.")
                driver.close()
                driver.quit()
                return True

        except Exception as e:
            print(f"Erreur lors de l'exécution", e)
            time.sleep(10000)
            driver.close()
            driver.quit()
            return False 

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)
        driver.close()
        driver.quit()
        return False  # Retourne False pour signaler une erreur


def retry_societe(file_path, sheet_name):
    """
    Fonction pour exécuter et relancer le traitement si une erreur se produit.
    """
    while True:  # Boucle infinie jusqu'à ce que le traitement soit terminé avec succès
        try:
            print(
                f"[INFO] Lancement du traitement : {file_path} - {sheet_name}")
            success = societe(file_path, sheet_name)

            if success:
                print(
                    f"[SUCCESS] Traitement terminé : {file_path} - {sheet_name}")
                break  # Succès, on sort de la boucle
            else:
                print(
                    f"[WARNING] Échec, relance dans 10s : {file_path} - {sheet_name}")
                time.sleep(10)

        except Exception as e:
            print(f"[ERROR] Erreur fatale : {e}")
            time.sleep(10)  # Attendre avant de réessayer


def launch_processes():
    """
    Fonction pour lancer les traitements en simultané.
    """
    # Liste pour stocker les processus

    for file_path, sheets in files_and_sheets:
        dep_number = os.path.basename(file_path).split('_')[1].split('.')[0]
        processes = []

        for sheet_name in sheets:
            print(f"Création d'un processus pour {file_path} - {sheet_name}")

            # Créer un processus pour chaque combinaison fichier/feuille
            process = Process(target=retry_societe,
                              args=(file_path, sheet_name))
            processes.append(process)
            process.start()  # Lancer le processus
            time.sleep(20)

        # Attendre que tous les processus soient terminés
        for process in processes:
            process.join()

        # Générer dynamiquement le nom du fichier fusionné
        departments_str = dep_number  # Ici, juste le département en cours
        directory = os.path.join(f"C:\\", "Users", user_name, "Desktop",
                                 "scrapping_aiscore", "societe", "Multi", f"DEPT_{dep_number}")
        print(f"Traitement du département {dep_number} terminé.")
        output_file = os.path.join(
            directory, f"news_dep_{departments_str}.xlsx")
        print(output_file)

        # Une fois tous les processus terminés, fusionner les fichierss
        merge_excel_files(output_file, dep_number, directory)

    print("Tous les départements ont été traités.")


def merge_excel_files(output_file, dep_number, directory):
    try:
        all_data = []

        # Parcourir les fichiers générés
        for file_path, parts in files_and_sheets:
            for part_name in parts:
                basename = f"DEPT_{dep_number}_{part_name}"
                individual_file = os.path.join(directory, f"{basename}.xlsx")
                if os.path.exists(individual_file):
                    df = pd.read_excel(individual_file)
                    all_data.append(df)

        # Fusionner toutes les données
        if all_data:
            merged_df = pd.concat(all_data, ignore_index=True)

            print(f"Suppression des doublons : {output_file}")
            merged_df = merged_df.drop_duplicates()  # Appliquer le résultat à merged_df

            merged_df.to_excel(output_file, index=False)
            print(f"Fichier fusionné créé : {output_file}")

        else:
            print("Aucun fichier à fusionner.")
    except Exception as e:
        print(f"Erreur lors de la fusion : {e}")


if __name__ == "__main__":
    print("Lancement des traitements en simultané...")
    launch_processes()
    print("Tous les traitements sont terminés.")
