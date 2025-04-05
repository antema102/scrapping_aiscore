import urllib.parse
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import os
from multiprocessing import Process, Lock
import time
import pandas as pd
import os
import random
import requests
from bs4 import BeautifulSoup
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Verrou global
files_and_sheets = []
lock = Lock()

# Récupérer l'utilisateur courant
user_name = os.getlogin()

for dep in range(59, 60):  # Départements de 8 à 12
    dep_formatted = str(dep).zfill(2)
    parts = [f"part_{j}" for j in range(1, 11)]  # Générer part_1 à part_6
    files_and_sheets.append(
        (f"C:/Users/{user_name}/Desktop/scrapping_aiscore/societe/Multi/DEPT/DEPT_{dep_formatted}.xlsx", parts)
    )

# Fonction pour charger les éléments déjà traités depuis un fichier spécifique
def load_processed_elements(filename):
    if os.path.exists(filename):
        with open(filename, 'r', encoding="utf-8") as file:
            return set(file.read().splitlines())
    return set()

# Fonction pour enregistrer les éléments traités dans un fichier spécifique
def save_processed_element(element_sirene, element_name, filename):
    with open(filename, 'a', encoding="utf-8") as file:
        file.write(f"{element_sirene} {element_name}\n")


def societe(file_path, sheets):
    try:
        # Proxy avec authentification (nom d'utilisateur et mot de passe)
        proxy = {
            'http': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
            'https': 'http://antema103.gmail.com:9yucvu@gate2.proxyfuel.com:2000',
        }

        # Liste de User-Agents
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.113 Safari/537.36",
            "Mozilla/5.0 (Linux; Android 10; Pixel 3 XL Build/QP1A.190711.020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Mobile Safari/537.36"
        ]

        user_agent = random.choice(user_agents)

        # Configuration des headers avec le User-Agent
        headers = {
            "User-Agent": user_agent
        }

        processed_text = os.path.splitext(os.path.basename(file_path))[0]
        number = processed_text.split("_")[-1]
        directory = os.path.join(f"DEPT_{number}")
        processed_filename = os.path.join(
            directory, f"{processed_text}_{sheets}.txt")
        new_file_path = os.path.join(
            directory, f"{processed_text}_{sheets}.xlsx")

        processed_elements = load_processed_elements(processed_filename)

        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Le dossier {directory} a été créé.")
        else:
            print(f"Le dossier {directory} existe déjà.")

        # Charger ou créer le fichier Excel
        if os.path.exists(new_file_path):
            workbook = load_workbook(new_file_path,)
            print(f"Le fichier {new_file_path} existe déjà.")
        else:
            workbook = load_workbook(file_path)
            print(f"Le fichier {new_file_path} a été créé.")
            # Parcourir toutes les feuilles et supprimer celles qui ne sont pas 'sheets'
            for feuille in workbook.sheetnames:
                if feuille != sheets:  # Si ce n'est pas l'onglet à garder
                    ws = workbook[feuille]
                    workbook.remove(ws)

        worksheet_name = sheets  # Nom de la feuille à garder dans le fichier Excel
        worksheet = workbook[worksheet_name]

        try:
            total_elements = worksheet.max_row  # Total des éléments dans la source de données
            processed_count = 0       # Compteur des éléments déjà traités
            # si ignoer alors code est for i, row in enumerate(ws[1:], start=2):
            for i, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
                name_company = row[1]
                code_postal = row[3]
                commune = row[4]   # Nom entreprise
                # Convertit en chaîne de caractères
                sirene_number = str(row[0])
                # Prend les 4 derniers chiffres
                last_four_digits_sirene = sirene_number[-4:]
                str_comparaison = f'{last_four_digits_sirene} {name_company}'
                if str_comparaison in processed_elements:
                    processed_count += 1
                    continue

                base_url = 'https://html.duckduckgo.com/html?q='
                query = f'site:www.societe.com {name_company} {code_postal} {commune}'
                encoded_query = urllib.parse.quote_plus(query)
                url = base_url + encoded_query

                max_retries = 4
                retry_delay = 5

                for attempt in range(max_retries):
                    try:
                        response = requests.get(
                            url, headers=headers, proxies=proxy, timeout=5, verify=False)

                        urllib3.disable_warnings(
                            urllib3.exceptions.InsecureRequestWarning)

                        if response.status_code == 200:

                            found_match = False
                            soup = BeautifulSoup(response.text, 'html.parser')
                            urls = soup.select(
                                '.result__extras__url a.result__url')

                            for item in urls:
                                try:
                                    href = item['href']
                                    sirene = item.text.split('-')[-1]
                                    siren_last = sirene.split('.')[0]
                                    last_four_digits = str(siren_last)[-4:]

                                    if last_four_digits == last_four_digits_sirene:
                                        found_match = True
                                        max_societe = 4
                                        retry_societe = 5
                                        # Essayer plusieurs fois en cas d'échec de la requête
                                        for attemp_societe in range(max_societe):

                                            try:
                                                new_response = requests.get(
                                                    href, headers=headers, proxies=proxy, timeout=10, verify=False)  # Timeout ajouté
                                                new_response.raise_for_status()  # Vérifier les erreurs HTTP
                                                # Si la requête réussit, traiter la réponse
                                                new_soup = BeautifulSoup(
                                                    new_response.text, 'html.parser')
                                                li = new_soup.select(
                                                    '.co-resume > ul > li')

                                                for item in li:
                                                    try:

                                                        span_text = item.select_one(
                                                            'span.ui-label').text.strip()
                                                        if span_text == 'ADRESSE':
                                                            span_adresse = item.select_one(
                                                                'span:nth-child(2) > a').text.strip()
                                                            span_adresse_parts = span_adresse.split(
                                                                ',')
                                                            span_adresse_str = span_adresse_parts[0] if len(
                                                                span_adresse_parts) > 0 else ''

                                                        if span_text == 'SIREN':
                                                            sirene_result = item.select_one(
                                                                'span:nth-child(2)').text.strip().replace(" ", "")

                                                    except Exception as e:
                                                        print(
                                                            'Erreur récupération du sirene et adresse:', e)

                                                worksheet.cell(
                                                    row=i, column=1, value=sirene_result)
                                                worksheet.cell(
                                                    row=i, column=3, value=span_adresse_str)

                                                print(
                                                    f"Sirène trouvé : noms {name_company} numero {sirene_result} addresse {span_adresse_str}  ligne {i}")
                                                workbook.save(
                                                    new_file_path)

                                                break

                                            except requests.exceptions.RequestException as e:

                                                print(
                                                    f"Tentative {attemp_societe + 1} échouée : {e}")
                                                if attemp_societe < max_societe - 1:
                                                    print(
                                                        f"Réessayer dans {retry_societe} secondes...")
                                                    time.sleep(retry_societe)
                                                else:
                                                    print(
                                                        f"Échec après {max_retries} tentatives.")
                                        break

                                except Exception as e:
                                    print(
                                        f"Erreur lors de la récupération du SIRENE: {e}")

                            # Si aucun match n'a été trouvé après la boucle
                            if not found_match:
                                print(
                                    f"Aucun sirene trouvé pour le nom  {name_company} code postal {code_postal} comune {commune}  ligne {i}")

                            processed_elements.add(
                                f"{last_four_digits_sirene} {name_company}")
                            save_processed_element(
                                last_four_digits_sirene, name_company, processed_filename)
                            processed_count += 1

                            break

                        else:
                            print(
                                f"⚠️ Statut {response.status_code}  pour le noms societes {name_company} != 200, tentative {attempt + 1}/{max_retries}")

                            if attempt < max_retries - 1:
                                time.sleep(retry_delay)

                    except Exception as e:
                        print("Captcha", e)
                        return False

                time.sleep(random.uniform(1, 5))

            # Vérifiez si tous les éléments ont été traités
            if processed_count >= total_elements:
                print("Tous les éléments ont été traités.")
                print("Script arrêté car aucune correspondance n'a été trouvée.")
                return True

        except Exception as e:
            print(f"Erreur lors de l'exécution", e)
            return False  # Retourne False pour signaler une erreur

    except Exception as e:
        print(f"Erreur lors de l'exécution _1", e)
        return False  # Retourne False pour signaler une erreur


def retry_societe(file_path, sheet_name):
    """
    Fonction pour exécuter et relancer le traitement si une erreur se produit.
    """
    while True:  # Boucle infinie jusqu'à ce que le traitement soit terminé avec succès
        try:
            print(
                f"[INFO] Lancement du traitement : {file_path} - {sheet_name}")
            success = societe(file_path, sheet_name)

            if success:
                print(
                    f"[SUCCESS] Traitement terminé : {file_path} - {sheet_name}")
                break  # Succès, on sort de la boucle
            else:
                print(
                    f"[WARNING] Échec, relance dans 10s : {file_path} - {sheet_name}")
                time.sleep(10)

        except Exception as e:
            print(f"[ERROR] Erreur fatale : {e}")
            time.sleep(10)  # Attendre avant de réessayer


def launch_processes():
    """
    Fonction pour lancer les traitements en simultané.
    """
    # Liste pour stocker les processus

    for file_path, sheets in files_and_sheets:
        dep_number = os.path.basename(file_path).split('_')[1].split('.')[0]
        processes = []

        for sheet_name in sheets:
            print(f"Création d'un processus pour {file_path} - {sheet_name}")

            # Créer un processus pour chaque combinaison fichier/feuille
            process = Process(target=retry_societe,
                              args=(file_path, sheet_name))
            processes.append(process)
            process.start()  # Lancer le processus
            time.sleep(20)

        # Attendre que tous les processus soient terminés
        for process in processes:
            process.join()

        # Générer dynamiquement le nom du fichier fusionné
        departments_str = dep_number  # Ici, juste le département en cours
        directory = os.path.join(f"C:\\", "Users", user_name, "Desktop",
                                 "scrapping_aiscore", "societe", "Multi", f"DEPT_{dep_number}")
        print(f"Traitement du département {dep_number} terminé.")
        output_file = os.path.join(
            directory, f"news_dep_{departments_str}.xlsx")
        print(output_file)

        # Une fois tous les processus terminés, fusionner les fichierss
        merge_excel_files(output_file, dep_number, directory)

    print("Tous les départements ont été traités.")


def merge_excel_files(output_file, dep_number, directory):
    try:
        all_data = []
        # Parcourir les fichiers générés
        for file_path, parts in files_and_sheets:
            for part_name in parts:
                basename = f"DEPT_{dep_number}_{part_name}"
                individual_file = os.path.join(directory, f"{basename}.xlsx")
                if os.path.exists(individual_file):
                    df = pd.read_excel(individual_file)
                    all_data.append(df)

        # Fusionner toutes les données
        if all_data:
            merged_df = pd.concat(all_data, ignore_index=True)
            print(f"Suppresion des doublons : {output_file}")
            merged_df.drop_duplicates()
            merged_df.to_excel(output_file, index=False)
            print(f"Fichier fusionné créé : {output_file}")

        else:
            print("Aucun fichier à fusionner.")
    except Exception as e:
        print("error merge ")


if __name__ == "__main__":
    print("Lancement des traitements en simultané...")
    launch_processes()
    print("Tous les traitements sont terminés.")
